"""
input_rka_widget.py — Halaman Input Rencana Kerja Anggaran (RKA)
"""
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QFrame,
    QComboBox, QLineEdit, QTableWidget, QTableWidgetItem, QHeaderView,
    QAbstractItemView, QStyledItemDelegate, QListView, QDialog, QGraphicsDropShadowEffect,
    QFileDialog, QMessageBox
)
from PySide6.QtCore import Qt, Signal
from PySide6.QtGui import QColor, QCursor, QFont, QIntValidator
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from core.db_manager import load_rka_record, get_connection
from ui.toast_notification import ToastManager
from ui.confirm_popup import ConfirmPopup
from ui.rka_form_popup import FormInputRKA
from ui.custom_dropdown import CustomDropdown
from ui.multi_select_dropdown import MultiSelectDropdown
import datetime

TABLE_ROWS = [
    {"label": "Dana Pihak Ketiga", "is_header": True, "cols": ["dpk_tabungan", "dpk_giro", "dpk_deposito"]},
    {"label": "Tabungan", "is_header": False, "cols": ["dpk_tabungan"]},
    {"label": "Giro", "is_header": False, "cols": ["dpk_giro"]},
    {"label": "Deposito", "is_header": False, "cols": ["dpk_deposito"]},
    {"label": "CASA", "is_header": False, "cols": ["dpk_tabungan", "dpk_giro"]},
    
    {"label": "DPK Korporasi", "is_header": True, "cols": ["korp_giro", "korp_deposito"]},
    {"label": "Giro", "is_header": False, "cols": ["korp_giro"]},
    {"label": "Deposito", "is_header": False, "cols": ["korp_deposito"]},
    
    {"label": "Pinjaman", "is_header": True, "cols": ["pinj_mikro", "pinj_small", "pinj_konsumer"]},
    {"label": "Mikro", "is_header": False, "cols": ["pinj_mikro"]},
    {"label": "Small", "is_header": False, "cols": ["pinj_small"]},
    {"label": "Konsumer", "is_header": False, "cols": ["pinj_konsumer"]},
    
    {"label": "SML", "is_header": True, "cols": ["sml_mikro", "sml_small", "sml_konsumer"]},
    {"label": "Mikro", "is_header": False, "cols": ["sml_mikro"]},
    {"label": "Small", "is_header": False, "cols": ["sml_small"]},
    {"label": "Konsumer", "is_header": False, "cols": ["sml_konsumer"]},
    
    {"label": "NPL", "is_header": True, "cols": ["npl_mikro", "npl_small", "npl_konsumer"]},
    {"label": "Mikro", "is_header": False, "cols": ["npl_mikro"]},
    {"label": "Small", "is_header": False, "cols": ["npl_small"]},
    {"label": "Konsumer", "is_header": False, "cols": ["npl_konsumer"]},
    
    {"label": "Recovery. EC", "is_header": True, "cols": ["rec_mikro", "rec_small", "rec_konsumer"]},
    {"label": "Mikro", "is_header": False, "cols": ["rec_mikro"]},
    {"label": "Small", "is_header": False, "cols": ["rec_small"]},
    {"label": "Konsumer", "is_header": False, "cols": ["rec_konsumer"]}
]

DEFAULT_KCS = [
    "KC Jakarta Tanah Abang",
    "KC Jakarta Krekot",
    "KC Jakarta Veteran",
    "KC Jakarta Roxi",
    "KC Jakarta Gunung Sahari",
    "KC Jakarta Mangga Dua",
    "KC Jakarta Kemayoran"
]

MONTHS_MAP = {
    "jan": "Januari", "feb": "Februari", "mar": "Maret", "apr": "April",
    "mei": "Mei", "jun": "Juni", "jul": "Juli", "agu": "Agustus",
    "sep": "September", "okt": "Oktober", "nov": "November", "des": "Desember"
}
MONTHS = list(MONTHS_MAP.keys())

COMBOBOX_STYLE = """
QComboBox {
    border: 1px solid #CBD5E1;
    border-radius: 6px;
    padding: 6px 12px;
    background: #FFFFFF;
    color: #334155;
    font-size: 13px;
}
QComboBox:focus {
    border: 1px solid #2563EB;
}
QComboBox QAbstractItemView {
    border: 1px solid #E2E8F0;
    border-radius: 6px;
    background-color: #FFFFFF;
    selection-background-color: #3B82F6;
    selection-color: #FFFFFF;
    color: #475569;
    outline: none;
    padding: 4px;
}
QComboBox QAbstractItemView::item {
    min-height: 32px;
    border-radius: 4px;
    padding: 4px 8px;
}
"""

class NumericDelegate(QStyledItemDelegate):
    def createEditor(self, parent, option, index):
        editor = QLineEdit(parent)
        return editor

class RKAFilterPopup(QDialog):
    def __init__(self, current_periode, current_kc, current_kats, current_subs, parent=None):
        super().__init__(parent)
        self.setWindowFlags(Qt.WindowType.Popup | Qt.WindowType.FramelessWindowHint | Qt.WindowType.NoDropShadowWindowHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        
        self.init_ui(current_periode, current_kc, current_kats, current_subs)
        
    def init_ui(self, p, kc, k, s):
        main_lay = QVBoxLayout(self)
        main_lay.setContentsMargins(10, 10, 10, 10)
        
        container = QFrame(self)
        container.setStyleSheet("QFrame#container { background: #FFFFFF; border: 1px solid #E2E8F0; border-radius: 16px; }")
        container.setObjectName("container")
        
        shadow = QGraphicsDropShadowEffect(self)
        shadow.setBlurRadius(20)
        shadow.setColor(QColor(0, 0, 0, 30))
        shadow.setOffset(0, 5)
        container.setGraphicsEffect(shadow)
        
        lay = QVBoxLayout(container)
        lay.setContentsMargins(24, 24, 24, 24)
        lay.setSpacing(12)
        
        title = QLabel("Filter Data RKA")
        title.setStyleSheet("font-size: 16px; font-weight: bold; color: #0F172A; background: transparent;")
        lay.addWidget(title)
        lay.addSpacing(8)
        
        lbl_p = QLabel("Periode")
        lbl_p.setStyleSheet("font-size: 12px; font-weight: bold; color: #64748B; background: transparent;")
        self.cb_periode = MultiSelectDropdown(["2026", "2027"])
        self.cb_periode.setCurrentSelections(p)
        
        lbl_kc = QLabel("Kantor Cabang")
        lbl_kc.setStyleSheet("font-size: 12px; font-weight: bold; color: #64748B; background: transparent;")
        self.cb_kc = MultiSelectDropdown(DEFAULT_KCS)
        self.cb_kc.setCurrentSelections(kc)
        
        lbl_k = QLabel("Mata Anggaran")
        lbl_k.setStyleSheet("font-size: 12px; font-weight: bold; color: #64748B; background: transparent;")
        main_kats = []
        for r in TABLE_ROWS:
            if r["is_header"]: main_kats.append(r["label"])
        self.cb_kategori = MultiSelectDropdown(main_kats)
        self.cb_kategori.setCurrentSelections(k)
        
        lbl_s = QLabel("Rincian")
        lbl_s.setStyleSheet("font-size: 12px; font-weight: bold; color: #64748B; background: transparent;")
        sub_kats = []
        for r in TABLE_ROWS:
            if not r["is_header"]: sub_kats.append(r["label"])
        self.cb_sub_kategori = MultiSelectDropdown(sub_kats)
        self.cb_sub_kategori.setCurrentSelections(s)
        
        lay.addWidget(lbl_p)
        lay.addWidget(self.cb_periode)
        lay.addSpacing(4)
        lay.addWidget(lbl_kc)
        lay.addWidget(self.cb_kc)
        lay.addSpacing(4)
        lay.addWidget(lbl_k)
        lay.addWidget(self.cb_kategori)
        lay.addSpacing(4)
        lay.addWidget(lbl_s)
        lay.addWidget(self.cb_sub_kategori)
        
        btn_lay = QHBoxLayout()
        btn_lay.setSpacing(12)
        
        btn_cancel = QPushButton("Batal")
        btn_cancel.setFixedHeight(36)
        btn_cancel.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        btn_cancel.setStyleSheet("QPushButton { background: #F1F5F9; color: #475569; border: none; border-radius: 18px; font-weight: 600; padding: 0 16px; } QPushButton:hover { background: #E2E8F0; }")
        btn_cancel.clicked.connect(self.reject)
        
        btn_apply = QPushButton("Terapkan Filter")
        btn_apply.setFixedHeight(36)
        btn_apply.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        btn_apply.setStyleSheet("QPushButton { background: #2563EB; color: #FFFFFF; border: none; border-radius: 18px; font-weight: 600; padding: 0 16px; } QPushButton:hover { background: #1D4ED8; }")
        btn_apply.clicked.connect(self.accept)
        
        btn_lay.addWidget(btn_cancel)
        btn_lay.addWidget(btn_apply)
        
        lay.addSpacing(16)
        lay.addLayout(btn_lay)
        
        main_lay.addWidget(container)
        self.setFixedWidth(320)

class InputRKAWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.is_editing = False
        current_year = str(datetime.datetime.now().year)
        self.current_periode = [current_year] if current_year in ["2026", "2027"] else ["2026"]
        self.current_kcs = ["Semua"]
        self.current_kats = ["Semua"]
        self.current_subs = ["Semua"]
        self.rka_data = []

        self.init_ui()
        self.load_data()

    def init_ui(self):
        main_lay = QVBoxLayout(self)
        main_lay.setContentsMargins(32, 32, 32, 32)
        main_lay.setSpacing(24)
        
        # ── HEADER ──
        hdr_lay = QHBoxLayout()
        hdr_left = QVBoxLayout()
        hdr_left.setSpacing(4)
        
        title = QLabel("Rencana Kerja Anggaran")
        title.setStyleSheet("font-size: 24px; font-weight: 700; color: #0F172A;")
        sub = QLabel("Target nominal per Kantor Cabang untuk setiap bulan (Jan–Des). Nilai dalam Juta Rupiah.")
        sub.setStyleSheet("font-size: 13px; color: #64748B;")
        
        hdr_left.addWidget(title)
        hdr_left.addWidget(sub)
        
        hdr_right = QHBoxLayout()
        hdr_right.setSpacing(12)
        
        self.btn_export = QPushButton("Export")
        self.btn_export.setFixedHeight(36)
        self.btn_export.setStyleSheet("QPushButton { background: #FFFFFF; border: 1px solid #E2E8F0; border-radius: 18px; color: #475569; padding: 0 20px; font-weight: 600; } QPushButton:hover { background: #F8FAFC; }")
        self.btn_export.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_export.clicked.connect(self.export_to_excel)
        
        self.btn_edit = QPushButton("Edit RKA")
        self.btn_edit.setFixedHeight(36)
        self.btn_edit.setStyleSheet("QPushButton { background: #2563EB; border: none; border-radius: 18px; color: #FFFFFF; padding: 0 20px; font-weight: 600; } QPushButton:hover { background: #1D4ED8; }")
        self.btn_edit.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_edit.clicked.connect(self.open_form)
        
        hdr_right.addWidget(self.btn_export)
        hdr_right.addWidget(self.btn_edit)
        
        hdr_lay.addLayout(hdr_left)
        hdr_lay.addStretch()
        hdr_lay.addLayout(hdr_right)
        
        # ── FILTER ROW ──
        flt_lay = QHBoxLayout()
        flt_lay.setSpacing(12)
        
        self.btn_filter = QPushButton(" Filter")
        self.btn_filter.setFixedHeight(36)
        self.btn_filter.setStyleSheet("QPushButton { background: #2563EB; border: none; border-radius: 18px; color: #FFFFFF; padding: 0 20px; font-weight: 600; } QPushButton:hover { background: #1D4ED8; }")
        self.btn_filter.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_filter.clicked.connect(self.open_filter_popup)
        
        self.btn_clear_filter = QPushButton("Clear Filter")
        self.btn_clear_filter.setFixedHeight(36)
        self.btn_clear_filter.setStyleSheet("QPushButton { background: transparent; color: #EF4444; border: 1px solid #EF4444; border-radius: 18px; padding: 0 16px; font-weight: 600; } QPushButton:hover { background: #FEE2E2; }")
        self.btn_clear_filter.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_clear_filter.clicked.connect(self.clear_filter)
        self.btn_clear_filter.hide()
        
        flt_lay.addWidget(self.btn_filter)
        flt_lay.addWidget(self.btn_clear_filter)
        
        flt_lay.addStretch()
        
        self.txt_search = QLineEdit()
        self.txt_search.setPlaceholderText("Cari KC...")
        self.txt_search.setFixedSize(200, 36)
        self.txt_search.setStyleSheet("QLineEdit { border: 1px solid #E2E8F0; border-radius: 18px; padding: 0 16px; background: #FFFFFF; }")
        self.txt_search.textChanged.connect(self._filter_table)
        
        btn_reset = QPushButton("Reset")
        btn_reset.setFixedHeight(36)
        btn_reset.setStyleSheet("QPushButton { background: transparent; color: #64748B; border: none; font-weight: 600; } QPushButton:hover { color: #0F172A; }")
        btn_reset.clicked.connect(lambda: self.txt_search.clear())
        
        flt_lay.addWidget(self.txt_search)
        flt_lay.addWidget(btn_reset)
        
        # ── SUMMARY CARDS ──
        cards_lay = QHBoxLayout()
        cards_lay.setSpacing(16)
        
        self.card_total = self._create_summary_card("TOTAL RKA TABUNGAN 2026", "0", "Juta Rupiah")
        self.card_avg = self._create_summary_card("RATA-RATA PER BULAN", "0", "Juta Rupiah", "#F0FDF4", "#047857")
        self.card_kc = self._create_summary_card("JUMLAH KANTOR CABANG", "10", "KC aktif", "#FFF7ED", "#C2410C")
        
        cards_lay.addWidget(self.card_total)
        cards_lay.addWidget(self.card_avg)
        cards_lay.addWidget(self.card_kc)
        
        # ── TABLE SECTION ──
        tbl_container = QFrame()
        tbl_container.setStyleSheet("QFrame { background: #FFFFFF; border-radius: 12px; border: 1px solid #E2E8F0; }")
        tbl_lay = QVBoxLayout(tbl_container)
        tbl_lay.setContentsMargins(20, 20, 20, 20)
        tbl_lay.setSpacing(12)
        
        tbl_hdr = QHBoxLayout()
        self.lbl_tbl_title = QLabel("Matriks RKA — Tabungan (2026)")
        self.lbl_tbl_title.setStyleSheet("font-size: 14px; font-weight: 700; color: #0F172A; border: none;")
        
        self.lbl_status = QLabel("REKAPITULASI")
        self.lbl_status.setStyleSheet("font-size: 10px; font-weight: 700; color: #047857; background: #D1FAE5; border-radius: 4px; padding: 4px 8px; border: none;")
        
        tbl_hdr.addWidget(self.lbl_tbl_title)
        tbl_hdr.addWidget(self.lbl_status)
        tbl_hdr.addStretch()
        
        self.tbl = QTableWidget()
        self.tbl.setColumnCount(15)
        headers = ["Kantor Cabang", "Jan", "Feb", "Mar", "Apr", "Mei", "Jun", "Jul", "Agu", "Sep", "Okt", "Nov", "Des", "Total", "Aksi"]
        self.tbl.setHorizontalHeaderLabels(headers)
        self.tbl.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.tbl.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self.tbl.horizontalHeader().setSectionResizeMode(14, QHeaderView.ResizeMode.Fixed)
        self.tbl.setColumnWidth(14, 80)
        self.tbl.verticalHeader().setVisible(False)
        self.tbl.setStyleSheet("""
            QTableWidget { border: none; gridline-color: #F1F5F9; color: #334155; }
            QHeaderView::section { background: #F8FAFC; color: #64748B; font-weight: 600; font-size: 12px; border: none; border-bottom: 1px solid #E2E8F0; padding: 8px; }
            QTableWidget::item { padding: 4px 8px; }
        """)
        self.tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        
        tbl_lay.addLayout(tbl_hdr)
        tbl_lay.addWidget(self.tbl)
        
        main_lay.addLayout(hdr_lay)
        main_lay.addLayout(flt_lay)
        main_lay.addLayout(cards_lay)
        main_lay.addWidget(tbl_container)

    def _create_tab_btn(self, text, active):
        btn = QPushButton(text)
        btn.setFixedHeight(36)
        btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        if active:
            btn.setStyleSheet("QPushButton { background: #2563EB; color: #FFFFFF; border: none; border-radius: 6px; padding: 0 16px; font-weight: 600; }")
        else:
            btn.setStyleSheet("QPushButton { background: #F1F5F9; color: #475569; border: none; border-radius: 6px; padding: 0 16px; font-weight: 600; } QPushButton:hover { background: #E2E8F0; }")
        btn.clicked.connect(lambda: self._on_tab_clicked(text))
        return btn
        
    def _create_summary_card(self, title, val, sub, bg="#F0F9FF", color="#0369A1"):
        frm = QFrame()
        frm.setStyleSheet(f"QFrame {{ background: {bg}; border-radius: 8px; border: none; }}")
        lay = QVBoxLayout(frm)
        lay.setContentsMargins(20, 20, 20, 20)
        lay.setSpacing(4)
        
        lbl_t = QLabel(title)
        lbl_t.setStyleSheet("font-size: 11px; font-weight: 700; color: #64748B; border: none;")
        
        lbl_v = QLabel(val)
        lbl_v.setObjectName("val")
        lbl_v.setStyleSheet(f"font-size: 28px; font-weight: 800; color: {color}; border: none;")
        
        lbl_s = QLabel(sub)
        lbl_s.setStyleSheet("font-size: 12px; color: #64748B; border: none;")
        
        lay.addWidget(lbl_t)
        lay.addWidget(lbl_v)
        lay.addWidget(lbl_s)
        return frm

    def open_form(self):
        target_kc = self.current_kcs[0] if self.current_kcs and "Semua" not in self.current_kcs else "KC Jakarta Tanah Abang"
        target_tahun = self.current_periode[0] if self.current_periode and "Semua" not in self.current_periode else "2026"
        
        form = FormInputRKA(target_kc, target_tahun, self.window())
        form.data_saved.connect(self.load_data)
        form.exec()
            
    def open_filter_popup(self):
        popup = RKAFilterPopup(self.current_periode, self.current_kategori, self.current_sub, self)
        
        # Position popup near the filter button
        pos = self.btn_filter.mapToGlobal(self.btn_filter.rect().bottomLeft())
        popup.move(pos.x(), pos.y() + 5)
        
        if popup.exec():
            self.current_periode = popup.cb_periode.currentText()
            self.current_kategori = popup.cb_kategori.currentText()
            self.current_sub = popup.cb_sub_kategori.currentText()
            
            if (self.current_periode != "2026" or 
                self.current_kategori != "Dana Pihak Ketiga" or 
                self.current_sub != "Total DPK"):
                self.btn_clear_filter.show()
                self.btn_filter.setStyleSheet("QPushButton { background: #F8FAFC; border: 1px solid #CBD5E1; border-radius: 18px; color: #334155; padding: 0 20px; font-weight: 600; } QPushButton:hover { background: #F1F5F9; }")
            else:
                self.btn_clear_filter.hide()
                self.btn_filter.setStyleSheet("QPushButton { background: #2563EB; border: none; border-radius: 18px; color: #FFFFFF; padding: 0 20px; font-weight: 600; } QPushButton:hover { background: #1D4ED8; }")
                
            self.load_data()

    def open_filter_popup(self):
        popup = RKAFilterPopup(self.current_periode, self.current_kcs, self.current_kats, self.current_subs, self)
        
        # Position popup near the filter button
        pos = self.btn_filter.mapToGlobal(self.btn_filter.rect().bottomLeft())
        popup.move(pos.x(), pos.y() + 5)
        
        if popup.exec():
            self.current_periode = popup.cb_periode.currentSelections()
            self.current_kcs = popup.cb_kc.currentSelections()
            self.current_kats = popup.cb_kategori.currentSelections()
            self.current_subs = popup.cb_sub_kategori.currentSelections()
            
            if (self.current_kcs != ["Semua"] or 
                self.current_kats != ["Semua"] or 
                self.current_subs != ["Semua"]):
                self.btn_clear_filter.show()
                self.btn_filter.setStyleSheet("QPushButton { background: #F8FAFC; border: 1px solid #CBD5E1; border-radius: 18px; color: #334155; padding: 0 20px; font-weight: 600; } QPushButton:hover { background: #F1F5F9; }")
            else:
                self.btn_clear_filter.hide()
                self.btn_filter.setStyleSheet("QPushButton { background: #2563EB; border: none; border-radius: 18px; color: #FFFFFF; padding: 0 20px; font-weight: 600; } QPushButton:hover { background: #1D4ED8; }")
                
            self.load_data()

    def clear_filter(self):
        current_year = str(datetime.datetime.now().year)
        self.current_periode = [current_year] if current_year in ["2026", "2027"] else ["2026"]
        self.current_kcs = ["Semua"]
        self.current_kats = ["Semua"]
        self.current_subs = ["Semua"]
        self.btn_clear_filter.hide()
        self.btn_filter.setStyleSheet("QPushButton { background: #2563EB; border: none; border-radius: 18px; color: #FFFFFF; padding: 0 20px; font-weight: 600; } QPushButton:hover { background: #1D4ED8; }")
        self.load_data()

    def load_data(self, _=None):
        target_kcs = DEFAULT_KCS if "Semua" in self.current_kcs else self.current_kcs
        target_periods = ["2026", "2027"] if "Semua" in self.current_periode else self.current_periode
        
        allowed_main = []
        if "Semua" in self.current_kats:
            for r in TABLE_ROWS:
                if r["is_header"]: allowed_main.append(r["label"])
        else:
            allowed_main = self.current_kats
            
        allowed_subs = []
        if "Semua" not in self.current_subs:
            allowed_subs = self.current_subs
            
        target_rows = []
        current_main = None
        for row in TABLE_ROWS:
            if row["is_header"]:
                current_main = row["label"]
            
            if current_main in allowed_main:
                if row["is_header"] or "Semua" in self.current_subs or row["label"] in allowed_subs:
                    target_rows.append(row)
        
        self.rka_data = []
        import sqlite3
        conn = get_connection()
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        for t in target_periods:
            tahun = int(t)
            for kc in target_kcs:
                cursor.execute(f"SELECT * FROM rka_data_v2 WHERE tahun = ? AND kc = ?", (tahun, kc))
                rows = cursor.fetchall()
                
                data_by_month = {m: None for m in range(1, 13)}
                for r in rows:
                    month_idx = 1
                    for short_m, long_m in MONTHS_MAP.items():
                        if long_m == r['bulan']:
                            data_by_month[month_idx] = dict(r)
                            break
                        month_idx += 1
                
                block_rows = []
                for row_cfg in target_rows:
                    row_data = {"label": row_cfg["label"], "is_header": row_cfg["is_header"], "months": [], "total": 0}
                    row_total = 0
                    for m in range(1, 13):
                        m_data = data_by_month[m]
                        val = 0
                        val_str = ""
                        if m_data:
                            if row_cfg.get("is_pct"):
                                val_str = str(m_data.get(row_cfg["cols"][0], "0,00"))
                            else:
                                for col in row_cfg["cols"]:
                                    val += m_data.get(col, 0)
                        
                        if row_cfg.get("is_pct"):
                            row_data["months"].append(val_str)
                        else:
                            row_data["months"].append(val)
                            row_total += val
                            
                    if row_cfg.get("is_pct"):
                        row_data["total"] = "0,00"
                    else:
                        row_data["total"] = row_total
                    block_rows.append(row_data)
                    
                self.rka_data.append({"kc": kc, "tahun": tahun, "rows": block_rows})
                
        conn.close()
        
        title_text = "Matriks RKA Keseluruhan" if len(target_kcs) == len(DEFAULT_KCS) else f"Matriks RKA ({len(target_kcs)} Cabang)"
        self.lbl_tbl_title.setText(title_text)
        self._populate_table()

    def export_to_excel(self):
        if not self.rka_data:
            ToastManager.show(self.window(), "Tidak ada data untuk di-export.", "warning")
            return
            
        tahun_str = "-".join(self.current_periode) if "Semua" not in self.current_periode else "Semua"
        default_filename = f"Dashboard RKA AH Gunsar Tahun {tahun_str}.xlsx"
            
        file_path, _ = QFileDialog.getSaveFileName(self, "Export RKA", default_filename, "Excel Files (*.xlsx)")
        if not file_path:
            return
            
        if not file_path.endswith('.xlsx'):
            file_path += '.xlsx'
            
        try:
            wb = Workbook()
            
            # Styles
            header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            align_center = Alignment(horizontal="center", vertical="center")
            align_right = Alignment(horizontal="right", vertical="center")
            align_left = Alignment(horizontal="left", vertical="center")
            
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # Headers
            headers = ["Mata Anggaran", "Jan", "Feb", "Mar", "Apr", "Mei", "Jun", "Jul", "Agu", "Sep", "Okt", "Nov", "Des", "Total"]
            
            for idx, block in enumerate(self.rka_data):
                kc_name = block['kc']
                safe_title = kc_name[:31]
                if not safe_title: safe_title = f"Sheet{idx+1}"
                
                if idx == 0:
                    ws = wb.active
                    ws.title = safe_title
                else:
                    ws = wb.create_sheet(title=safe_title)
                    
                for col_idx, col_name in enumerate(headers, start=1):
                    cell = ws.cell(row=1, column=col_idx, value=col_name)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = align_center
                    cell.border = thin_border
                    
                ws.column_dimensions['A'].width = 30
                for col_idx in range(2, 15):
                    ws.column_dimensions[chr(64+col_idx)].width = 15
                    
                current_row = 2
                
                # KC Header Row
                cell = ws.cell(row=current_row, column=1, value=f"{block['kc']} ({block['tahun']})")
                cell.font = Font(bold=True, color="1E3A8A", size=12)
                cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
                cell.alignment = align_left
                
                # Fill rest of row with color
                for c in range(1, 15):
                    ws.cell(row=current_row, column=c).fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
                    ws.cell(row=current_row, column=c).border = thin_border
                    
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=14)
                current_row += 1
                
                for data in block["rows"]:
                    label = data["label"]
                    is_header = data["is_header"]
                    
                    label_cell = ws.cell(row=current_row, column=1, value=label)
                    label_cell.border = thin_border
                    
                    font_bold = Font(bold=True)
                    if is_header:
                        label_cell.font = font_bold
                        label_cell.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
                    else:
                        label_cell.alignment = Alignment(indent=2)
                        
                    for i in range(12):
                        val = data["months"][i]
                        val_cell = ws.cell(row=current_row, column=i+2, value=val)
                        val_cell.alignment = align_right
                        val_cell.border = thin_border
                        if is_header:
                            val_cell.font = font_bold
                            val_cell.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
                        if not isinstance(val, str):
                            val_cell.number_format = '#,##0'
                            
                    tot = data["total"]
                    tot_cell = ws.cell(row=current_row, column=14, value=tot)
                    tot_cell.alignment = align_right
                    tot_cell.border = thin_border
                    tot_cell.font = font_bold
                    if is_header:
                        tot_cell.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
                    if not isinstance(tot, str):
                        tot_cell.number_format = '#,##0'
                        
                    current_row += 1
                
            wb.save(file_path)
            ToastManager.show(self.window(), "Export Excel berhasil!", "success")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Gagal export Excel: {e}")

    def _filter_table(self):
        self._populate_table()

    def _populate_table(self):
        self.tbl.blockSignals(True)
        self.tbl.clearSpans()
        
        search_txt = self.txt_search.text().lower()
        
        total_rows = 0
        for block in self.rka_data:
            if search_txt and search_txt not in block["kc"].lower():
                continue
            total_rows += 1 # KC Title
            total_rows += len(block["rows"])
            total_rows += 1 # Empty spacing
            
        self.tbl.setRowCount(total_rows)
        self.tbl.setColumnCount(14)
        headers = ["Mata Anggaran", "Jan", "Feb", "Mar", "Apr", "Mei", "Jun", "Jul", "Agu", "Sep", "Okt", "Nov", "Des", "Total"]
        self.tbl.setHorizontalHeaderLabels(headers)
        self.tbl.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.tbl.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        
        row_idx = 0
        for block in self.rka_data:
            if search_txt and search_txt not in block["kc"].lower():
                continue
                
            # KC Title Row
            it_kc = QTableWidgetItem(f"   {block['kc']} ({block['tahun']})")
            it_kc.setFlags(Qt.ItemFlag.ItemIsEnabled)
            font_kc = QFont()
            font_kc.setBold(True)
            font_kc.setPointSize(12)
            it_kc.setFont(font_kc)
            it_kc.setBackground(QColor("#DBEAFE"))
            it_kc.setForeground(QColor("#1E3A8A"))
            self.tbl.setItem(row_idx, 0, it_kc)
            for c in range(1, 14):
                it = QTableWidgetItem()
                it.setFlags(Qt.ItemFlag.ItemIsEnabled)
                it.setBackground(QColor("#DBEAFE"))
                self.tbl.setItem(row_idx, c, it)
            self.tbl.setSpan(row_idx, 0, 1, 14)
            row_idx += 1
            
            for data in block["rows"]:
                label = data["label"]
                is_header = data["is_header"]
                
                # Label item
                it_lbl = QTableWidgetItem(label)
                it_lbl.setFlags(Qt.ItemFlag.ItemIsEnabled)
                font = QFont()
                if is_header:
                    font.setBold(True)
                else:
                    it_lbl.setText("    " + label) # Indent
                it_lbl.setFont(font)
                
                if is_header:
                    it_lbl.setBackground(QColor("#F8FAFC"))
                
                self.tbl.setItem(row_idx, 0, it_lbl)
                
                for i in range(12):
                    val = data["months"][i]
                    if isinstance(val, str):
                        it_val = QTableWidgetItem(val)
                    else:
                        it_val = QTableWidgetItem(f"{val:,}".replace(",", "."))
                        
                    it_val.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                    it_val.setFlags(Qt.ItemFlag.ItemIsEnabled)
                    if is_header:
                        it_val.setFont(font)
                        it_val.setBackground(QColor("#F8FAFC"))
                    self.tbl.setItem(row_idx, i+1, it_val)
                    
                # Total
                tot = data["total"]
                if isinstance(tot, str):
                    it_tot = QTableWidgetItem(tot)
                else:
                    it_tot = QTableWidgetItem(f"{tot:,}".replace(",", "."))
                    
                it_tot.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                it_tot.setFlags(Qt.ItemFlag.ItemIsEnabled)
                font_bold = QFont()
                font_bold.setBold(True)
                it_tot.setFont(font_bold)
                if is_header:
                    it_tot.setBackground(QColor("#F8FAFC"))
                self.tbl.setItem(row_idx, 13, it_tot)
                row_idx += 1
                
            # Spacing
            for c in range(14):
                it = QTableWidgetItem()
                it.setFlags(Qt.ItemFlag.NoItemFlags)
                self.tbl.setItem(row_idx, c, it)
            self.tbl.setSpan(row_idx, 0, 1, 14)
            self.tbl.setRowHeight(row_idx, 20)
            row_idx += 1
            
        dpk_total = 0
        if self.rka_data and self.rka_data[0]["rows"]:
            dpk_total = self.rka_data[0]["rows"][0]["total"]
            
        val_lbl = self.card_total.findChild(QLabel, "val")
        if val_lbl: val_lbl.setText(f"{dpk_total:,}".replace(",", "."))
        
        avg_lbl = self.card_avg.findChild(QLabel, "val")
        if avg_lbl: avg_lbl.setText(f"{int(dpk_total/12):,}".replace(",", ".") if dpk_total else "0")
        
        kc_lbl = self.card_kc.findChild(QLabel, "val")
        if kc_lbl: kc_lbl.setText(str(len(self.rka_data)))
        
        self.tbl.blockSignals(False)
