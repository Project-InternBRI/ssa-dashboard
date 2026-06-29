"""
history_preview_dialog.py — Popup preview tabel dari file Excel riwayat.

QDialog fullscreen-like (90% lebar × 85% tinggi).
Tab per sheet Excel. QTableWidget per sheet.
Filter KC + Export Sheet Ini di bagian bawah.
"""
import os
import shutil
from pathlib import Path

from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel,
    QPushButton, QFrame, QTabWidget, QTableWidget,
    QTableWidgetItem, QHeaderView, QFileDialog,
    QWidget, QSizePolicy, QApplication, QScrollArea,
    QLineEdit,
)
from PySide6.QtCore import Qt, Signal
from PySide6.QtGui import QCursor, QColor, QBrush, QFont

try:
    import openpyxl
    from openpyxl.styles import PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from ui.toast_notification import ToastManager
    HAS_TOAST = True
except ImportError:
    HAS_TOAST = False


class HistoryPreviewDialog(QDialog):
    """
    Dialog preview tabel untuk satu entri riwayat.
    Membaca file Excel dengan openpyxl dan menampilkan per sheet.
    """

    def __init__(self, excel_path: str, entry: dict, parent=None):
        super().__init__(
            parent,
            Qt.WindowType.Dialog | Qt.WindowType.WindowCloseButtonHint)
        self.setWindowTitle("Preview Riwayat Dashboard")
        self.setModal(True)

        self._excel_path = excel_path
        self._entry = entry
        self._sheets: dict[str, list[list]] = {}  # sheet_name → rows of cells

        # Ukuran dialog: 90% lebar × 85% tinggi
        if parent:
            pg = parent.geometry()
            w = int(pg.width() * 0.90)
            h = int(pg.height() * 0.85)
        else:
            screen = QApplication.primaryScreen().availableGeometry()
            w = int(screen.width() * 0.90)
            h = int(screen.height() * 0.85)

        self.resize(w, h)
        self.setMinimumSize(900, 600)

        self._build_ui()
        self._load_excel()

    # ─── BUILD UI ─────────────────────────────────────────────────
    def _build_ui(self):
        self.setStyleSheet("""
            QDialog { background: #FFFFFF; }
            QTabWidget::pane { border: none; background: #FFFFFF; }
            QTabBar::tab {
                background: transparent; color: #64748B;
                padding: 8px 16px; border: none;
                border-bottom: 2px solid transparent;
                font-size: 12px;
            }
            QTabBar::tab:selected {
                color: #2563EB; font-weight: 600;
                border-bottom: 2px solid #2563EB;
            }
            QTabBar::tab:hover:!selected { color: #334155; }
        """)

        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ── Header ────────────────────────────────────────────────
        header = QFrame()
        header.setFixedHeight(56)
        header.setStyleSheet("""
            QFrame {
                background: #1E293B;
                border: none;
            }
        """)
        hdr_lay = QHBoxLayout(header)
        hdr_lay.setContentsMargins(24, 0, 16, 0)
        hdr_lay.setSpacing(12)

        tgl = self._entry.get("tanggal_data", "")
        lbl_title = QLabel(f"Dashboard AH Gunsar  —  {tgl}")
        lbl_title.setStyleSheet(
            "color: #F1F5F9; font-size: 14px; font-weight: 600; background: transparent;")
        hdr_lay.addWidget(lbl_title)
        hdr_lay.addStretch()

        btn_close = QPushButton("Tutup  X")
        btn_close.setFixedHeight(32)
        btn_close.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        btn_close.setStyleSheet("""
            QPushButton {
                background: rgba(255,255,255,0.08); color: #94A3B8;
                border: 1px solid rgba(255,255,255,0.12); border-radius: 6px;
                padding: 0 14px; font-size: 12px;
            }
            QPushButton:hover {
                background: rgba(239,68,68,0.15); color: #FCA5A5;
                border-color: rgba(239,68,68,0.3);
            }
        """)
        btn_close.clicked.connect(self.close)
        hdr_lay.addWidget(btn_close)

        # ── Loading indicator ─────────────────────────────────────
        self._lbl_loading = QLabel("Memuat data Excel...")
        self._lbl_loading.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._lbl_loading.setStyleSheet(
            "font-size: 14px; color: #64748B; padding: 40px;")

        # ── Tab widget ────────────────────────────────────────────
        self._tabs = QTabWidget()
        self._tabs.setDocumentMode(True)
        self._tabs.hide()

        # ── Footer ───────────────────────────────────────────────
        footer = QFrame()
        footer.setFixedHeight(52)
        footer.setStyleSheet("""
            QFrame {
                background: #F8FAFC;
                border-top: 1px solid #E2E8F0;
            }
        """)
        foot_lay = QHBoxLayout(footer)
        foot_lay.setContentsMargins(20, 0, 20, 0)
        foot_lay.setSpacing(12)

        # Info KC
        kc = self._entry.get('jumlah_kc', 0)
        wkt = self._entry.get('waktu_proses', '—')
        lbl_info = QLabel(
            f"{kc} KC  ·  {self._entry.get('jumlah_periode', 0)} periode  ·  "
            f"Diproses: {self._entry.get('tanggal_proses', '')[:16]}")
        lbl_info.setStyleSheet("font-size: 12px; color: #64748B; background: transparent;")
        foot_lay.addWidget(lbl_info)
        foot_lay.addStretch()

        # Tombol Export Sheet Ini
        self._btn_export_sheet = QPushButton("Export Sheet Ini")
        self._btn_export_sheet.setFixedHeight(34)
        self._btn_export_sheet.setCursor(
            QCursor(Qt.CursorShape.PointingHandCursor))
        self._btn_export_sheet.setStyleSheet("""
            QPushButton {
                background: #10B981; color: white;
                border: none; border-radius: 6px;
                padding: 0 16px; font-weight: 600; font-size: 12px;
            }
            QPushButton:hover { background: #059669; }
        """)
        self._btn_export_sheet.clicked.connect(self._export_current_sheet)

        # Tombol Export Semua
        btn_export_all = QPushButton("Export Semua")
        btn_export_all.setFixedHeight(34)
        btn_export_all.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        btn_export_all.setStyleSheet("""
            QPushButton {
                background: #2563EB; color: #FFFFFF;
                border: none; border-radius: 6px;
                padding: 0 16px; font-weight: 500; font-size: 13px;
            }
            QPushButton:hover { background: #1D4ED8; }
        """)
        btn_export_all.clicked.connect(self._export_all)

        foot_lay.addWidget(self._btn_export_sheet)
        foot_lay.addWidget(btn_export_all)

        # Susun layout
        root.addWidget(header)
        root.addWidget(self._lbl_loading, 1)
        root.addWidget(self._tabs, 1)
        root.addWidget(footer)

    # ─── LOAD EXCEL ───────────────────────────────────────────────
    def _load_excel(self):
        """Baca file Excel dan populate QTabWidget."""
        if not HAS_OPENPYXL:
            self._lbl_loading.setText(
                "openpyxl tidak tersedia. Instal dengan: pip install openpyxl")
            return

        path = Path(self._excel_path)
        if not path.exists():
            self._lbl_loading.setText(
                f"File tidak ditemukan:\n{self._excel_path}")
            return

        try:
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        except Exception as e:
            self._lbl_loading.setText(f"Gagal membuka file Excel:\n{e}")
            return

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Baca semua baris
            all_rows = []
            for row in ws.iter_rows(values_only=True):
                all_rows.append(list(row))

            if not all_rows:
                continue

            self._sheets[sheet_name] = all_rows
            tab_widget = self._build_table(all_rows)
            self._tabs.addTab(tab_widget, sheet_name)

        wb.close()

        if self._tabs.count() > 0:
            self._lbl_loading.hide()
            self._tabs.show()
        else:
            self._lbl_loading.setText("File Excel tidak memiliki sheet data.")

    # ─── BUILD TABLE ──────────────────────────────────────────────
    def _build_table(self, all_rows: list[list]) -> QWidget:
        """Buat QWidget berisi QTableWidget dari baris Excel."""
        if not all_rows:
            w = QWidget()
            QVBoxLayout(w).addWidget(QLabel("Sheet kosong."))
            return w

        # Hitung kolom maks
        max_cols = max((len(r) for r in all_rows), default=1)
        n_rows = len(all_rows)

        tbl = QTableWidget(n_rows, max_cols)
        tbl.setAlternatingRowColors(False)
        tbl.setShowGrid(True)
        tbl.setGridStyle(Qt.PenStyle.SolidLine)
        tbl.verticalHeader().hide()
        tbl.horizontalHeader().setStretchLastSection(False)
        tbl.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Interactive)
        tbl.setHorizontalScrollMode(QTableWidget.ScrollMode.ScrollPerPixel)
        tbl.setVerticalScrollMode(QTableWidget.ScrollMode.ScrollPerPixel)
        tbl.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        tbl.setSelectionMode(QTableWidget.SelectionMode.NoSelection)
        tbl.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        tbl.setStyleSheet("""
            QTableWidget::item:hover { background-color: transparent; }
            QTableWidget::item:selected { background-color: transparent; }
        """)

        # Warna style yang mirip dengan Excel Export
        header_bg = QColor("#1E3A8A")  # Biru tua elegan
        header_fg = QColor("#FFFFFF")
        alt_a = QColor("#FFFFFF")
        alt_b = QColor("#F8FAFC")
        bold_bg = QColor("#EFF6FF")    # Biru sangat muda untuk row total
        bold_labels = ['Dana Pihak Ketiga', 'CASA', 'DPK Korporasi', 'Pinjaman', 'SML', 'SML %', 'NPL', 'NPL %']

        bold_font = QFont()
        bold_font.setBold(True)
        
        italic_font = QFont()
        italic_font.setItalic(True)

        # Populate
        for r_idx, row_data in enumerate(all_rows):
            for c_idx in range(max_cols):
                val = row_data[c_idx] if c_idx < len(row_data) else None

                # Format value ke Rupiah / format angka Indonesia
                if val is None:
                    text = ""
                elif isinstance(val, (int, float)):
                    if isinstance(val, float) and val != int(val):
                        s = f"{val:,.2f}"
                    else:
                        s = f"{int(val):,}"
                    text = s.replace(',', 'X').replace('.', ',').replace('X', '.')
                else:
                    text = str(val)

                item = QTableWidgetItem(text)
                item.setFlags(Qt.ItemFlag.ItemIsEnabled)  # Hapus ItemIsSelectable agar hover tidak mengubah warna

                if r_idx == 0:
                    item.setBackground(QBrush(QColor("#17365D")))
                    item.setForeground(QBrush(QColor("white")))
                    item.setFont(bold_font)
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
                elif r_idx == 1:
                    item.setBackground(QBrush(QColor("#F0F8FF")))
                    item.setForeground(QBrush(QColor("#1E293B")))
                    item.setFont(italic_font)
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
                elif r_idx == 2:
                    item.setBackground(QBrush(QColor("#17365D")))
                    item.setForeground(QBrush(QColor("white")))
                    item.setFont(bold_font)
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
                elif r_idx == 3:
                    item.setBackground(QBrush(QColor("#2563EB")))
                    item.setForeground(QBrush(QColor("white")))
                    item.setFont(bold_font)
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
                else:
                    is_bold = (row_data[0] in bold_labels) if len(row_data) > 0 else False
                    
                    if is_bold:
                        item.setBackground(QBrush(bold_bg))
                        item.setFont(bold_font)
                    else:
                        bg = alt_a if r_idx % 2 == 0 else alt_b
                        item.setBackground(QBrush(bg))
                    
                    item.setForeground(QBrush(QColor("#1E293B")))

                    if c_idx == 0:
                        item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                    else:
                        item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)

                tbl.setItem(r_idx, c_idx, item)

        # --- Lakukan Pergabungan Sel (Merge/Span) ---
        if n_rows > 0:
            tbl.setSpan(0, 0, 1, max_cols)  # Judul
        if n_rows > 1:
            tbl.setSpan(1, 0, 1, max_cols)  # Sub-judul
            
        if n_rows > 3:
            tbl.setSpan(2, 0, 2, 1)  # Mata Anggaran merge ke bawah
            
            # Deteksi span untuk Posisi, RKA, dll di baris ke-2 (index 2)
            row2 = all_rows[2]
            c = 1
            while c < max_cols:
                val = row2[c]
                if val not in (None, ""):
                    nxt = c + 1
                    while nxt < max_cols and row2[nxt] in (None, ""):
                        nxt += 1
                    span_len = nxt - c
                    if span_len > 1:
                        tbl.setSpan(2, c, 1, span_len)
                    c = nxt
                else:
                    c += 1

        # Kolom pertama lebih lebar
        tbl.setColumnWidth(0, 200)
        for c in range(1, max_cols):
            tbl.setColumnWidth(c, 110)

        # Sembunyikan header angka (1, 2, 3...) yang membuat jelek
        tbl.horizontalHeader().hide()
        
        # Sembunyikan baris 1 ("Diekspor pada...") sesuai permintaan
        if n_rows > 1:
            tbl.setRowHidden(1, True)

        wrapper = QWidget()
        lay = QVBoxLayout(wrapper)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.addWidget(tbl)
        return wrapper

    # ─── EXPORT ───────────────────────────────────────────────────
    def _export_current_sheet(self):
        """Export sheet yang sedang aktif ke file Excel baru."""
        if not self._tabs.count():
            return

        sheet_name = self._tabs.tabText(self._tabs.currentIndex())
        dest, _ = QFileDialog.getSaveFileName(
            self, "Export Sheet",
            f"Dashboard_{sheet_name}.xlsx",
            "Excel Files (*.xlsx)")
        if not dest:
            return

        try:
            import shutil
            import openpyxl
            
            src = str(self._excel_path)
            # Salin file asli agar format, warna, merge, dan grafik terbawa
            shutil.copy2(src, dest)
            
            # Buka file hasil salinan dan hapus sheet yang tidak dipilih
            wb_new = openpyxl.load_workbook(dest)
            for sn in wb_new.sheetnames:
                if sn != sheet_name:
                    del wb_new[sn]
                    
            wb_new.save(dest)
            if HAS_TOAST:
                ToastManager.show(
                    self.window(),
                    f"Sheet '{sheet_name}' berhasil diekspor.", "success")
        except Exception as e:
            if HAS_TOAST:
                ToastManager.show(self.window(), f"Gagal ekspor: {e}", "error")

    def _export_all(self):
        """Copy file Excel asli ke lokasi baru."""
        src = Path(self._excel_path)
        if not src.exists():
            if HAS_TOAST:
                ToastManager.show(
                    self.window(), "File tidak ditemukan.", "error")
            return

        dest, _ = QFileDialog.getSaveFileName(
            self, "Simpan File Excel", src.name, "Excel Files (*.xlsx)")
        if not dest:
            return

        try:
            shutil.copy2(str(src), dest)
            if HAS_TOAST:
                ToastManager.show(
                    self.window(), "File berhasil diekspor.", "success")
        except Exception as e:
            if HAS_TOAST:
                ToastManager.show(
                    self.window(), f"Gagal ekspor: {e}", "error")
