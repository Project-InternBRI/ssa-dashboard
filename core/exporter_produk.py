import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color.lstrip("#"))

def export_monitoring_produk_to_excel(data_dict: dict, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Monitoring Produk"

    # Define styles
    header_fill = _fill("00B0F0")
    orange_fill = _fill("FFC000")
    gray_fill = _fill("D9D9D9")
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    num_format = '#,##0;[Red](#,##0);"-"'
    pct_format = '0.00%;[Red](0.00%);"-"'

    # Filter out non-KC keys
    skip_keys = {'__stats__', '__uker_data__'}
    
    # KCs should be sorted, but let's keep the order in data_dict
    kc_names = [k for k in data_dict.keys() if k not in skip_keys and k != 'Total AH Gunsar']
    kc_names.sort() # sort alphabetically
    
    # Make sure 'Total AH Gunsar' is added at the end
    if 'Total AH Gunsar' in data_dict:
        kc_names.append('Total AH Gunsar')
        
    if not kc_names:
        raise ValueError("Tidak ada data KC untuk di-export.")

    # Get periode_list and metadata from the first KC to build headers
    first_kc = kc_names[0]
    first_data = data_dict[first_kc]
    periode_list = first_data.get('periode_list', [])
    
    rows = first_data.get('rows', [])
    meta = next((r for r in rows if r.get('row_type') == '__metadata__'), None)
    terbaru_dt = None
    if meta:
        terbaru_dt = meta['periode_refs']['terbaru']

    from core.file_reader import BULAN_SINGKAT
    from core.processor import format_label
    
    rka_label = "RKA"
    pencapaian_label = "Pencapaian RKA"
    if terbaru_dt:
        rka_label = f"RKA {BULAN_SINGKAT[terbaru_dt.month]} {str(terbaru_dt.year)[-2:]}"
        pencapaian_label = f"Pencapaian RKA\n{BULAN_SINGKAT[terbaru_dt.month]} {str(terbaru_dt.year)[-2:]}"

    # Define the 5 tables we want to generate
    tables_config = [
        ("Tabungan", "Tabungan", "data"),
        ("Giro", "Giro", "data"),
        ("Deposito", "Deposito", "data"),
        ("Total Casa", "CASA", "bold"),
        ("Total DPK", "Dana Pihak Ketiga", "header_value")
    ]

    current_row = 1

    for title, row_label, row_type in tables_config:
        # Title of the table
        ws.cell(row=current_row, column=1, value=title).font = bold_font
        current_row += 1
        
        # Headers
        c = 1
        cell = ws.cell(row=current_row, column=c, value="Branch Office")
        cell.font = bold_font; cell.alignment = center_align; cell.border = thin_border; cell.fill = header_fill
        c += 1
        
        for p in periode_list:
            cell = ws.cell(row=current_row, column=c, value=p)
            cell.font = bold_font; cell.alignment = center_align; cell.border = thin_border; cell.fill = header_fill
            c += 1
            
        cell = ws.cell(row=current_row, column=c, value=rka_label)
        cell.font = bold_font; cell.alignment = center_align; cell.border = thin_border; cell.fill = orange_fill
        c += 1
        
        cell = ws.cell(row=current_row, column=c, value="MTD")
        cell.font = bold_font; cell.alignment = center_align; cell.border = thin_border; cell.fill = orange_fill
        c += 1
        
        cell = ws.cell(row=current_row, column=c, value="YTD")
        cell.font = bold_font; cell.alignment = center_align; cell.border = thin_border; cell.fill = orange_fill
        c += 1
        
        cell = ws.cell(row=current_row, column=c, value="DTD")
        cell.font = bold_font; cell.alignment = center_align; cell.border = thin_border; cell.fill = orange_fill
        c += 1
        
        cell = ws.cell(row=current_row, column=c, value=pencapaian_label)
        cell.font = bold_font; cell.alignment = center_align; cell.border = thin_border; cell.fill = orange_fill
        ws.row_dimensions[current_row].height = 30
        c += 1
        
        current_row += 1
        
        # Data Rows
        for kc in kc_names:
            kc_rows = data_dict[kc].get('rows', [])
            # Find the target row
            target_r = next((r for r in kc_rows if r.get('label') == row_label and r.get('row_type') == row_type), None)
            
            is_total = (kc == 'Total AH Gunsar')
            display_name = 'Total Area Head Gunung Sahari' if is_total else kc
            
            c = 1
            cell = ws.cell(row=current_row, column=c, value=display_name)
            cell.border = thin_border
            if is_total:
                cell.font = bold_font
                cell.fill = gray_fill
            c += 1
            
            if target_r:
                vals = target_r.get('values', {})
                for p in periode_list:
                    val = vals.get(p)
                    cell = ws.cell(row=current_row, column=c, value=val)
                    cell.number_format = num_format
                    cell.border = thin_border
                    if is_total:
                        cell.font = bold_font; cell.fill = gray_fill
                    c += 1
                    
                terbaru_label = format_label(terbaru_dt) if terbaru_dt else periode_list[-1]
                val_terbaru = vals.get(terbaru_label)
                
                pencapaian = target_r.get('pencapaian_rka')
                rka_val = None
                if pencapaian and pencapaian != 0 and val_terbaru:
                    rka_val = val_terbaru / pencapaian
                    
                cell = ws.cell(row=current_row, column=c, value=rka_val)
                cell.number_format = num_format
                cell.border = thin_border
                if is_total:
                    cell.font = bold_font; cell.fill = gray_fill
                c += 1
                
                # MTD, YTD, DTD (already in Juta)
                for metric in ['mtd', 'ytd', 'dtd']:
                    val = target_r.get(metric)
                    cell = ws.cell(row=current_row, column=c, value=val)
                    cell.number_format = num_format
                    cell.border = thin_border
                    if is_total:
                        cell.font = bold_font; cell.fill = gray_fill
                    c += 1
                    
                # Pencapaian RKA
                cell = ws.cell(row=current_row, column=c, value=pencapaian)
                cell.number_format = pct_format
                cell.border = thin_border
                if is_total:
                    cell.font = bold_font; cell.fill = gray_fill
                c += 1
                
            else:
                # Fill empty row
                for _ in range(len(periode_list) + 5):
                    cell = ws.cell(row=current_row, column=c, value="-")
                    cell.border = thin_border
                    if is_total:
                        cell.fill = gray_fill
                    c += 1
            
            current_row += 1
            
        # Add empty row between tables
        current_row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 45
    for col_idx in range(2, 2 + len(periode_list) + 5):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15

    wb.save(output_path)

