"""
exporter.py — Export hasil proses ke format Excel template Dashboard AH Gunsar.

Struktur sheet:
  2 baris info (judul + tanggal export)
  2 baris header (grup + detail)
  N baris data (template FIXED)

Kolom:
  Mata Anggaran | POSISI (per periode) | RKA | Pencp RKA % | GROWTH (MTD|DTD|YOY|YTD)

Freeze panes: B5 (2 baris info + 2 baris header)
"""
from __future__ import annotations
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation

from core.file_reader import BULAN_PANJANG, BULAN_SINGKAT


# ────────────────────────────────────────────────────────────────────
# STYLE HELPERS
# ────────────────────────────────────────────────────────────────────
def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color.lstrip("#"))

def get_default_export_filename(data: dict, kc_name: str = "AH Gunsar") -> str:
    date_str = ""
    try:
        kc_data = data.get(kc_name)
        if not kc_data:
            kc_data = data.get("Total AH Gunsar")
        if not kc_data:
            kc_data = next(iter(data.values()))
            
        rows = kc_data.get("rows", [])
        meta_row = next((r for r in rows if r.get('row_type') == '__metadata__'), None)
        if meta_row:
            terbaru_dt = meta_row['periode_refs']['terbaru']
            from core.file_reader import BULAN_PANJANG
            if terbaru_dt:
                date_str = f" {terbaru_dt.day} {BULAN_PANJANG[terbaru_dt.month]} {terbaru_dt.year}"
    except Exception:
        pass
        
    base_name = f"Dashboard {kc_name}{date_str}.xlsx"
    if kc_name == "AH Gunsar":
        # Specific naming for Export All as requested
        base_name = f"Dashboard AH Gunsar{date_str}.xlsx"
        
    return base_name

def get_unique_path(base_path: str) -> str:
    import os
    if not os.path.exists(base_path):
        return base_path
        
    name, ext = os.path.splitext(base_path)
    counter = 1
    while True:
        new_path = f"{name} ({counter}){ext}"
        if not os.path.exists(new_path):
            return new_path
        counter += 1


def _font(bold=False, italic=False, size=10,
          color="1E293B", name="Arial") -> Font:
    return Font(name=name, bold=bold, italic=italic, size=size,
                color=color.lstrip("#"))


def _align(h="center", v="center", wrap=False, indent=0) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap,
                     indent=indent)


def _border(style="thin", color="E2E8F0") -> Border:
    s = Side(style=style, color=color.lstrip("#"))
    return Border(left=s, right=s, top=s, bottom=s)


def _border_bottom_medium() -> Border:
    thin = Side(style="thin", color="E2E8F0")
    med = Side(style="medium", color="1E3A5F")
    return Border(left=thin, right=thin, top=thin, bottom=med)


_NO_BORDER = Border()


# Warna
CLR_HEADER_BG     = "1E3A5F"
CLR_HEADER2_BG    = "2563EB"
CLR_HEADER_FONT   = "FFFFFF"
CLR_SECTION_BG    = "DBEAFE"    # Header seksi (biru muda)
CLR_BOLD_BG       = "EFF6FF"    # Bold label
CLR_SEPARATOR_BG  = "1E3A5F"    # Separator biru solid
CLR_ROW_A         = "FFFFFF"
CLR_ROW_B         = "F8FAFC"
CLR_RKA_BG        = "E2E8F0"    # Abu muda sesuai mockup
CLR_PENCAP_BG     = "E2E8F0"    # Abu muda sesuai mockup
CLR_NEG_FONT      = "DC2626"
CLR_POS_FONT      = "1E293B"
CLR_ZERO_FONT     = "94A3B8"
CLR_INFO_BG       = "1E3A5F"
CLR_INFO2_BG      = "EFF6FF"
# ────────────────────────────────────────────────────────────────────
# DASHBOARD HELPERS — DATA ACCESS
# ────────────────────────────────────────────────────────────────────

def get_valid_kc_list(data_dict):
    """Return list of valid KC keys, filtering out Total AH Gunsar and __stats__."""
    excluded = {"Total AH Gunsar", "__stats__"}
    valid_kc = []
    for key in data_dict.keys():
        if key in excluded:
            continue
        if key.startswith('_'):
            continue
        if not isinstance(data_dict[key], dict):
            continue
        if 'rows' not in data_dict[key]:
            continue
        valid_kc.append(key)
    return valid_kc


def get_val_helper(data_dict, kc, section, label, periode, key='values'):
    kc_data = data_dict.get(kc)
    if not kc_data or "rows" not in kc_data:
        return None
    curr_sec = ""
    for r in kc_data["rows"]:
        rt = r.get("row_type", "")
        lbl = r.get("label", "")
        if rt in ("header", "header_value") or (rt == "bold" and lbl in ("SML", "NPL", "Recovery. EC")):
            curr_sec = lbl
        if curr_sec == section and lbl == label:
            if key == 'values' and periode:
                val = r.get("values", {}).get(periode)
                return val if val is not None else None
            val = r.get(key)
            return val if val is not None else None
    return None


def get_periode_terbaru(data_dict):
    total_data = data_dict.get("Total AH Gunsar")
    if not total_data:
        return None
    periode_list = total_data.get("periode_list", [])
    return periode_list[-1] if periode_list else None


# ────────────────────────────────────────────────────────────────────
# MAIN DASHBOARD FUNCTION — COMPLETE REWRITE
# ────────────────────────────────────────────────────────────────────

def build_dashboard_visual(ws, data_dict, metadata=None):
    """Build dashboard with 8 bar charts (MTD+YTD) and time series line chart."""
    import os
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.chart.label import DataLabelList

    if metadata is None:
        metadata = {'tanggal_terbaru': '', 'jam': ''}

    kc_list = get_valid_kc_list(data_dict)
    periode_terbaru = get_periode_terbaru(data_dict)

    if not kc_list or not periode_terbaru:
        print("[ERROR] Data tidak valid — dashboard tidak dibuat")
        return

    periode_list = data_dict.get("Total AH Gunsar", {}).get("periode_list", [])

    def _get_mtd(kc, section, label):
        return get_val_helper(data_dict, kc, section, label, None, key='mtd') or 0

    def _get_ytd(kc, section, label):
        return get_val_helper(data_dict, kc, section, label, None, key='ytd') or 0

    def _get_val(kc, section, label, periode):
        return get_val_helper(data_dict, kc, section, label, periode) or 0

    metrics = []
    for kc in kc_list:
        metrics.append({
            'kc': kc,
            'mtd_dpk': (_get_mtd(kc, "Dana Pihak Ketiga", "Dana Pihak Ketiga") + _get_mtd(kc, "DPK Korporasi", "DPK Korporasi")) / 1000,
            'mtd_tab': _get_mtd(kc, "Dana Pihak Ketiga", "Tabungan") / 1000,
            'mtd_gir': (_get_mtd(kc, "Dana Pihak Ketiga", "Giro") + _get_mtd(kc, "DPK Korporasi", "Giro")) / 1000,
            'mtd_dep': (_get_mtd(kc, "Dana Pihak Ketiga", "Deposito") + _get_mtd(kc, "DPK Korporasi", "Deposito")) / 1000,
            'ytd_dpk': (_get_ytd(kc, "Dana Pihak Ketiga", "Dana Pihak Ketiga") + _get_ytd(kc, "DPK Korporasi", "DPK Korporasi")) / 1000,
            'ytd_tab': _get_ytd(kc, "Dana Pihak Ketiga", "Tabungan") / 1000,
            'ytd_gir': (_get_ytd(kc, "Dana Pihak Ketiga", "Giro") + _get_ytd(kc, "DPK Korporasi", "Giro")) / 1000,
            'ytd_dep': (_get_ytd(kc, "Dana Pihak Ketiga", "Deposito") + _get_ytd(kc, "DPK Korporasi", "Deposito")) / 1000
        })

    # KPI totals
    total_dpk = _get_val("Total AH Gunsar", "Dana Pihak Ketiga", "Dana Pihak Ketiga", periode_terbaru) + _get_val("Total AH Gunsar", "DPK Korporasi", "DPK Korporasi", periode_terbaru)
    total_tab = _get_val("Total AH Gunsar", "Dana Pihak Ketiga", "Tabungan", periode_terbaru)
    total_giro = _get_val("Total AH Gunsar", "Dana Pihak Ketiga", "Giro", periode_terbaru) + _get_val("Total AH Gunsar", "DPK Korporasi", "Giro", periode_terbaru)
    total_dep = _get_val("Total AH Gunsar", "Dana Pihak Ketiga", "Deposito", periode_terbaru) + _get_val("Total AH Gunsar", "DPK Korporasi", "Deposito", periode_terbaru)

    # ──────────────────────────────────────────────
    # STEP 2: PREPARE AND WRITE CHART SOURCE DATA
    # ──────────────────────────────────────────────
    DATA_START = 150
    DC = 30  # Column AD

    # We have 16 charts: 4 sections * 4 metrics. Each needs 2 columns (KC, Value)
    # 0-3: MTD Kenaikan (DPK, Tab, Gir, Dep) -> cols 30-37
    # 4-7: MTD Penurunan -> cols 38-45
    # 8-11: YTD Kenaikan -> cols 46-53
    # 12-15: YTD Penurunan -> cols 54-61

    chart_data_ranges = [] # list of (min_row, max_row, col_idx_kc, col_idx_val)
    col_offset = 0

    for metric_type in ['mtd', 'ytd']:
        for is_penurunan in [False, True]:
            for metric_key in ['dpk', 'tab', 'gir', 'dep']:
                full_key = f"{metric_type}_{metric_key}"
                
                # Filter and sort
                if not is_penurunan:
                    filtered = [(m['kc'], m[full_key]) for m in metrics if m[full_key] > 0]
                    filtered.sort(key=lambda x: x[1], reverse=True) # Highest positive first
                else:
                    filtered = [(m['kc'], m[full_key]) for m in metrics if m[full_key] < 0]
                    filtered.sort(key=lambda x: x[1]) # Lowest negative first (e.g. -5000 before -2000)

                # Take top 5
                top5 = filtered[:5]
                
                kc_col = DC + col_offset
                val_col = DC + col_offset + 1
                
                # Write to sheet
                for i, (k, v) in enumerate(top5):
                    ws.cell(row=DATA_START + i, column=kc_col).value = k
                    ws.cell(row=DATA_START + i, column=val_col).value = v
                
                max_row = DATA_START + len(top5) - 1 if top5 else DATA_START
                chart_data_ranges.append((DATA_START, max_row, kc_col, val_col))
                
                col_offset += 2

    # Time Series Data
    TS_START = DATA_START + 10
    for i, p in enumerate(periode_list):
        r = TS_START + i
        ws.cell(row=r, column=DC+32).value = str(p)
        ws.cell(row=r, column=DC+33).value = (_get_val("Total AH Gunsar", "Dana Pihak Ketiga", "Dana Pihak Ketiga", p) + _get_val("Total AH Gunsar", "DPK Korporasi", "DPK Korporasi", p)) / 1000
        ws.cell(row=r, column=DC+34).value = _get_val("Total AH Gunsar", "Dana Pihak Ketiga", "Tabungan", p) / 1000
        ws.cell(row=r, column=DC+35).value = (_get_val("Total AH Gunsar", "Dana Pihak Ketiga", "Giro", p) + _get_val("Total AH Gunsar", "DPK Korporasi", "Giro", p)) / 1000
        ws.cell(row=r, column=DC+36).value = (_get_val("Total AH Gunsar", "Dana Pihak Ketiga", "Deposito", p) + _get_val("Total AH Gunsar", "DPK Korporasi", "Deposito", p)) / 1000
    TS_END = TS_START + len(periode_list) - 1

    # ──────────────────────────────────────────────
    # STEP 3: COLUMN WIDTHS & ZOOM
    # ──────────────────────────────────────────────
    ws.sheet_view.zoomScale = 100
    ws.column_dimensions['A'].width = 2
    for col in 'BCDEFGHIJKLMNOPQRSTU':
        ws.column_dimensions[col].width = 11.5

    # ──────────────────────────────────────────────
    # STEP 4: HEADER (rows 1-3)
    # ──────────────────────────────────────────────
    ws.merge_cells('B1:U2')
    ws.row_dimensions[1].height = 60
    ws.row_dimensions[2].height = 40
    ws.row_dimensions[3].height = 20 # Spacer gap

    header_title = "DASHBOARD TABUNGAN, GIRO & DEPOSITO"

    from datetime import datetime
    tgl = metadata.get('tanggal_terbaru')
    if not tgl:
        tgl = str(periode_terbaru)
        try:
            if " " in tgl: tgl = tgl.split()[0]
            dt = datetime.strptime(tgl, '%Y-%m-%d')
            months = ["", "Jan", "Feb", "Mar", "Apr", "Mei", "Jun", "Jul", "Agu", "Sep", "Okt", "Nov", "Des"]
            tgl = f"{dt.day} {months[dt.month]} {dt.year}"
        except:
            pass
    jam = metadata.get('jam')
    jam_str = f" {jam}" if jam else ""
    header_subtitle = f"Data per {tgl}{jam_str} WIB"

    try:
        from PIL import Image, ImageDraw, ImageFont
        from openpyxl.drawing.image import Image as XLImage
        import tempfile
        import os
        import math

        W, H = 1900, 135 # Height matches rows 1+2 (100 pts ~= 135 px)
        
        bg_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'assets', 'background_header_excel_dashboard.png')
        if not os.path.exists(bg_path):
            bg_path = os.path.join('assets', 'background_header_excel_dashboard.png')

        if os.path.exists(bg_path):
            orig_bg = Image.open(bg_path).convert("RGBA")
            base_color = orig_bg.getpixel((0, 0))
            bg = Image.new("RGBA", (W, H), base_color)
            
            # Fit fully without cropping
            piece_h = H
            piece_w = int(orig_bg.width * (piece_h / orig_bg.height))
            # Just in case they wanted it a bit wider, we can stretch it slightly
            # But let's stick to strict aspect ratio to prevent cropping/distortion
            piece = orig_bg.resize((piece_w, piece_h), Image.Resampling.LANCZOS)
            
            mask = Image.new("L", (piece_w, piece_h), 255)
            fade_width = min(200, piece_w) # smooth transition area
            for x in range(fade_width):
                # Cosine wave for very smooth transition
                alpha = int(255 * (1 - math.cos(math.pi * x / fade_width)) / 2)
                for y in range(piece_h):
                    mask.putpixel((x, y), alpha)
                    
            piece.putalpha(mask)
            bg.paste(piece, (W - piece_w, 0), piece)
        else:
            bg = Image.new("RGBA", (W, H), "#002060")

        logo_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'assets', 'icons', 'bri_excel_logo.png')
        if not os.path.exists(logo_path):
            logo_path = os.path.join('assets', 'icons', 'bri_excel_logo.png')

        if os.path.exists(logo_path):
            orig_logo = Image.open(logo_path)
            has_alpha = 'A' in orig_logo.mode or ('transparency' in orig_logo.info)
            logo = orig_logo.convert("RGBA")
            
            if has_alpha:
                alpha = logo.split()[3]
                white_logo = Image.new("RGBA", logo.size, (255, 255, 255, 255))
                white_logo.putalpha(alpha)
                logo = white_logo
                
            lh = 65
            lw = int(logo.width * (lh / logo.height))
            logo = logo.resize((lw, lh), Image.Resampling.LANCZOS)
            bg.paste(logo, (50, (H - lh)//2), logo)
            text_x = 50 + lw + 40

            draw = ImageDraw.Draw(bg)
            draw.line([(text_x - 20, 25), (text_x - 20, H - 25)], fill="white", width=3)
        else:
            text_x = 50
            draw = ImageDraw.Draw(bg)

        try:
            font_title = ImageFont.truetype("/System/Library/Fonts/Helvetica.ttc", 40)
            font_sub = ImageFont.truetype("/System/Library/Fonts/Helvetica.ttc", 22)
        except:
            font_title = ImageFont.load_default()
            font_sub = ImageFont.load_default()

        draw.text((text_x, (H//2) - 45), header_title, font=font_title, fill="white")
        draw.text((text_x, (H//2) + 10), f"📅 {header_subtitle}", font=font_sub, fill="white")

        temp_img_path = os.path.join(tempfile.gettempdir(), 'header_generated.png')
        bg.save(temp_img_path)

        xl_img = XLImage(temp_img_path)
        xl_img.width = W
        xl_img.height = H
        ws.add_image(xl_img, 'B1')

    except Exception as e:
        print(f"[HEADER] Gagal generate header image: {e}")
        ws.merge_cells('B1:U1')
        h1 = ws['B1']
        h1.value = header_title
        h1.font = Font(name='Calibri', size=18, bold=True, color='FFFFFF')
        h1.fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
        h1.alignment = Alignment(horizontal='left', vertical='center', indent=2)

        ws.merge_cells('B2:U2')
        h2 = ws['B2']
        h2.value = header_subtitle
        h2.font = Font(name='Calibri', size=10, italic=True, color='FFFFFF')
        h2.fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
        h2.alignment = Alignment(horizontal='left', vertical='center', indent=2)


    # ──────────────────────────────────────────────
    # STEP 5: KPI CARDS (rows 4-8)
    # ──────────────────────────────────────────────
    thin_s = Side(style='thin', color='CBD5E1')
    card_border = Border(top=thin_s, bottom=thin_s, left=thin_s, right=thin_s)

    kpi_configs = [
        ("TOTAL DANA PIHAK KETIGA (DPK)", total_dpk, '000080', '000080', 'kpi_dpk.png', 'B'),
        ("TABUNGAN", total_tab, '000080', '000080', 'kpi_tab.png', 'G'),
        ("GIRO", total_giro, 'FF6600', '000080', 'kpi_gir.png', 'L'),
        ("DEPOSITO", total_dep, '006600', '006600', 'kpi_dep.png', 'Q')
    ]

    for lbl, val, title_color, val_color, icon_name, sc in kpi_configs:
        sc_i = column_index_from_string(sc)
        ec_i = sc_i + 4
        ec = get_column_letter(ec_i)
        
        # Merge for Icon (sc, rows 4-7)
        ws.merge_cells(start_row=4, start_column=sc_i, end_row=7, end_column=sc_i)
        
        # Attempt to insert icon
        try:
            from openpyxl.drawing.image import Image as XLImage
            img_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'assets', 'icons', icon_name)
            if not os.path.exists(img_path):
                img_path = os.path.join('assets', 'icons', icon_name)
            if os.path.exists(img_path):
                img = XLImage(img_path)
                img.width = 54
                img.height = 54
                ws.add_image(img, f'{sc}4')
        except Exception as e:
            print(f"[KPI] Gagal load icon {icon_name}: {e}")

        # Title: sc+1 row 4
        ws.merge_cells(start_row=4, start_column=sc_i+1, end_row=4, end_column=ec_i)
        c = ws.cell(row=4, column=sc_i+1)
        c.value = lbl
        c.font = Font(name='Calibri', size=10, bold=True, color=title_color)
        c.alignment = Alignment(horizontal='left', vertical='bottom', wrap_text=True)

        # Date: sc+1 row 5
        ws.merge_cells(start_row=5, start_column=sc_i+1, end_row=5, end_column=ec_i)
        c = ws.cell(row=5, column=sc_i+1)
        c.value = f"Per {periode_terbaru}"
        c.font = Font(name='Calibri', size=9, color='000080')
        c.alignment = Alignment(horizontal='left', vertical='top')

        # Value: sc+1 row 6
        ws.merge_cells(start_row=6, start_column=sc_i+1, end_row=6, end_column=ec_i)
        c = ws.cell(row=6, column=sc_i+1)
        c.value = val
        c.number_format = '#,##0'
        c.font = Font(name='Calibri', size=20, bold=True, color=val_color)
        c.alignment = Alignment(horizontal='left', vertical='center')

        # Rp: sc+1 row 7
        ws.merge_cells(start_row=7, start_column=sc_i+1, end_row=7, end_column=ec_i)
        c = ws.cell(row=7, column=sc_i+1)
        c.value = "Rp"
        c.font = Font(name='Calibri', size=9, bold=True, color=val_color)
        c.alignment = Alignment(horizontal='left', vertical='top')

        # Borders & Background
        for r_i in range(4, 8):
            for c_i in range(sc_i, ec_i + 1):
                cell = ws.cell(row=r_i, column=c_i)
                # Apply outer border only
                top_s = thin_s if r_i == 4 else Side(border_style=None)
                bottom_s = thin_s if r_i == 7 else Side(border_style=None)
                left_s = thin_s if c_i == sc_i else Side(border_style=None)
                right_s = thin_s if c_i == ec_i else Side(border_style=None)
                cell.border = Border(top=top_s, bottom=bottom_s, left=left_s, right=right_s)
                cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 14
    ws.row_dimensions[6].height = 26
    ws.row_dimensions[7].height = 14
    ws.row_dimensions[8].height = 10 # Spacer below cards

    # ──────────────────────────────────────────────
    # HELPER TO BUILD CHART SECTIONS
    # ──────────────────────────────────────────────
    def _build_chart_section(start_row, title, chart_index_offset, is_penurunan):
        ws.merge_cells(f'B{start_row}:U{start_row}')
        s = ws[f'B{start_row}']; s.value = title
        # Plain text formatting as requested, no background fill
        s.font = Font(name='Calibri', size=14, bold=True, color='000080')
        s.fill = PatternFill(fill_type=None)
        s.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[start_row].height = 24

        for lbl, sc, ec, color in [
            ("DANA PIHAK KETIGA\n(Rp Miliar)", 'B', 'F', '1E3A8A'),
            ("TABUNGAN\n(Rp Miliar)", 'G', 'K', '2563EB'),
            ("GIRO\n(Rp Miliar)", 'L', 'P', 'F97316'),
            ("DEPOSITO\n(Rp Miliar)", 'Q', 'U', '16A34A')
        ]:
            ws.merge_cells(f'{sc}{start_row+1}:{ec}{start_row+1}')
            c = ws[f'{sc}{start_row+1}']; c.value = lbl
            c.font = Font(name='Calibri', size=9, bold=True, color='FFFFFF')
            c.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[start_row+1].height = 28

        for idx, (anchor, color) in enumerate([('B', '1E3A8A'), ('G', '2563EB'), ('L', 'F97316'), ('Q', '16A34A')]):
            chart_data = chart_data_ranges[chart_index_offset + idx]
            min_r, max_r, kc_col, val_col = chart_data
            
            c = BarChart()
            c.type = "bar"
            c.width = 11.5; c.height = 7.5; c.gapWidth = 100; c.legend = None
            c.visible_cells_only = False
            
            if max_r >= min_r:
                c.add_data(Reference(ws, min_col=val_col, min_row=min_r, max_row=max_r))
                c.set_categories(Reference(ws, min_col=kc_col, min_row=min_r, max_row=max_r))
            
            c.y_axis.numFmt = '#,##0.##'
            c.x_axis.tickLblPos = "low" # Ensures KC names are on the far left
            
            if c.series:
                c.series[0].graphicalProperties.solidFill = color if not is_penurunan else '1E3A8A' # Use dark blue for all bars like screenshot? Actually use theme color or green/orange
                
                # Based on user screenshot, chart colors match the header colors!
                c.series[0].graphicalProperties.solidFill = color
                
                dl = DataLabelList()
                dl.showVal = True
                dl.position = "outEnd" # Label outside the end of the bar
                c.series[0].dLbls = dl
                
            ws.add_chart(c, f"{anchor}{start_row+2}")

    # ──────────────────────────────────────────────
    # BUILD ALL 4 SECTIONS
    # ──────────────────────────────────────────────
    # chart_index_offset: 0=MTD Ken, 4=MTD Pen, 8=YTD Ken, 12=YTD Pen
    _build_chart_section(9, f"KENAIKAN TERTINGGI MTD ({periode_terbaru})", 0, False)
    _build_chart_section(25, f"PENURUNAN TERTINGGI MTD ({periode_terbaru})", 4, True)
    _build_chart_section(41, f"KENAIKAN TERTINGGI YTD (1 JAN - {periode_terbaru})", 8, False)
    _build_chart_section(57, f"PENURUNAN TERTINGGI YTD (1 JAN - {periode_terbaru})", 12, True)

    # ──────────────────────────────────────────────
    # TIME SERIES SECTION (row 73-88)
    # ──────────────────────────────────────────────
    ws.merge_cells('B73:U73'); s = ws['B73']; s.value = "TREND PERKEMBANGAN DPK (Rp Miliar)"
    s.font = Font(name='Calibri', size=14, bold=True, color='000080'); s.fill = PatternFill(fill_type=None)
    s.alignment = Alignment(horizontal='left', vertical='center', indent=1); ws.row_dimensions[73].height = 24

    ts = LineChart(); ts.width = 34; ts.height = 10; ts.y_axis.title = "Rp Miliar"; ts.x_axis.title = "Periode"; ts.title = "Trend DPK, Tabungan, Giro & Deposito"
    ts.visible_cells_only = False
    for j, c in enumerate(['1E3A8A', '2563EB', 'F97316', '16A34A']):
        ts.add_data(Reference(ws, min_col=DC+33+j, min_row=TS_START, max_row=TS_END))
        if j < len(ts.series): ts.series[j].graphicalProperties.line.solidFill = c
    ts.set_categories(Reference(ws, min_col=DC+32, min_row=TS_START, max_row=TS_END))
    ts.y_axis.numFmt = '#,##0'; ws.add_chart(ts, "B74")

    # ──────────────────────────────────────────────
    # FOOTER & CLEANUP
    # ──────────────────────────────────────────────
    fr = 89
    ws.merge_cells(f'B{fr}:U{fr}'); fc = ws[f'B{fr}']; fc.value = "Catatan: Data bersifat rahasia dan hanya untuk penggunaan internal."
    fc.font = Font(name='Calibri', size=9, italic=True, color='FFFFFF'); fc.fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    fc.alignment = Alignment(horizontal='center', vertical='center'); ws.row_dimensions[fr].height = 18

    for col_idx in range(22, max(120, ws.max_column + 5)): ws.column_dimensions[get_column_letter(col_idx)].hidden = True
    ws.sheet_view.showGridLines = False

    print(f"[SUCCESS] Dashboard selesai — footer di baris {fr}")

# ────────────────────────────────────────────────────────────────────
# EXPORT UTAMA
# ────────────────────────────────────────────────────────────────────



def export_to_excel(data_dict: dict,
                    output_path: str,
                    tanggal_data: str = "") -> Path:
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    
    # ====== KODE BARU — TAMBAHKAN INI ======
    ws_dashboard = wb.active
    ws_dashboard.title = "Dashboard"
    
    total_ah_gunsar = data_dict.get("Total AH Gunsar", {})
    build_dashboard_visual(ws_dashboard, data_dict, metadata=None)
    # ====== AKHIR KODE BARU ======

    # Urutan sheet sesuai spesifikasi
    kc_order = [
        'Tanah Abang', 'Krekot', 'Veteran', 'Roxi',
        'Gunung Sahari', 'Mangga Dua', 'Kemayoran',
    ]

    for kc_name in kc_order:
        if kc_name in data_dict:
            kc_data = data_dict[kc_name]
            ws = wb.create_sheet(title=kc_name[:31])
            _write_sheet(ws, kc_name, kc_data)

    if "Total AH Gunsar" in data_dict:
        ws = wb.create_sheet(title="Total AH Gunsar")
        _write_sheet(ws, "Total AH Gunsar", data_dict["Total AH Gunsar"])

    try:
        wb.save(str(out))
    except Exception as e:
        raise RuntimeError(f"Gagal menyimpan file Excel: {e}") from e

    return out


# ────────────────────────────────────────────────────────────────────
# TULIS SATU SHEET
# ────────────────────────────────────────────────────────────────────
def _write_sheet(ws, kc_name: str, kc_data: dict) -> None:
    periode_list = kc_data.get("periode_list", [])
    rows_data    = kc_data.get("rows", [])
    mtd_label    = kc_data.get("mtd_label", "MTD")
    dtd_label    = kc_data.get("dtd_label", "DTD")
    yoy_label    = kc_data.get("yoy_label", "YOY")
    ytd_label    = kc_data.get("ytd_label", "YTD")

    n_periode = len(periode_list)

    # Kolom layout
    COL_MATA      = 1
    COL_POS_START = 2
    COL_POS_END   = max(COL_POS_START, COL_POS_START + n_periode - 1)
    COL_RKA_START = COL_POS_END + 1
    COL_RKA_END   = COL_RKA_START + 11
    COL_PENCAP    = COL_RKA_END + 1
    COL_MTD       = COL_PENCAP + 1
    COL_DTD       = COL_MTD + 1
    COL_YOY       = COL_DTD + 1
    COL_YTD       = COL_YOY + 1
    LAST_COL      = COL_YTD

    # ══════════════════════════════════════════════════════════════
    # BARIS 1-2: INFO EXPORT
    # ══════════════════════════════════════════════════════════════
    now = datetime.now()
    HARI_ID = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Sabtu', 'Minggu']
    hari_str = HARI_ID[now.weekday()]
    tgl_str = f"{now.day} {BULAN_PANJANG[now.month]} {now.year}"
    jam_str = now.strftime("%H:%M") + " WIB"
    export_info = f"Diekspor pada: {tgl_str} {hari_str}, {jam_str}"

    # Baris 1: Judul
    for ci in range(1, LAST_COL + 1):
        c = ws.cell(row=1, column=ci,
                    value=f"Dashboard SSA — Bank BRI AH Gunsar Jakarta Region ({kc_name})"
                    if ci == 1 else "")
        c.fill = _fill(CLR_INFO_BG)
        c.font = _font(bold=True, size=13, color=CLR_HEADER_FONT)
        c.alignment = _align(h="center", v="center")

    if LAST_COL > 1:
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=LAST_COL)
    ws.row_dimensions[1].height = 28

    # Baris 2: Tanggal export
    for ci in range(1, LAST_COL + 1):
        c = ws.cell(row=2, column=ci,
                    value=export_info if ci == 1 else "")
        c.fill = _fill(CLR_INFO2_BG)
        c.font = _font(italic=True, size=10, color="64748B")
        c.alignment = _align(h="center", v="center")

    if LAST_COL > 1:
        ws.merge_cells(start_row=2, start_column=1,
                       end_row=2, end_column=LAST_COL)
    ws.row_dimensions[2].height = 22

    # ══════════════════════════════════════════════════════════════
    # BARIS 3-4: HEADER TABEL
    # ══════════════════════════════════════════════════════════════
    HDR_ROW1 = 3
    HDR_ROW2 = 4

    header_fill = _fill(CLR_HEADER_BG)
    header_font = _font(bold=True, size=11, color=CLR_HEADER_FONT)
    header_align = _align(h="center", v="center", wrap=True)
    header_bdr = _border("medium", "1E3A5F")

    header2_fill = _fill(CLR_HEADER2_BG)
    header2_font = _font(bold=True, size=10, color=CLR_HEADER_FONT)
    header2_align = _align(h="center", v="center", wrap=True)

    # A3: "Mata Anggaran" (merge A3:A4)
    for r in [HDR_ROW1, HDR_ROW2]:
        c = ws.cell(row=r, column=COL_MATA,
                    value="Mata Anggaran" if r == HDR_ROW1 else "")
        c.fill = header_fill
        c.font = header_font
        c.alignment = header_align
        c.border = header_bdr

    # Posisi header grup (baris 3)
    for ci in range(COL_POS_START, COL_POS_END + 1):
        c = ws.cell(row=HDR_ROW1, column=ci,
                    value="Posisi" if ci == COL_POS_START else "")
        c.fill = header_fill
        c.font = header_font
        c.alignment = header_align
        c.border = header_bdr

    # Posisi detail (baris 4) — label periode
    for i, lbl in enumerate(periode_list):
        c = ws.cell(row=HDR_ROW2, column=COL_POS_START + i, value=lbl)
        c.fill = header2_fill
        c.font = header2_font
        c.alignment = header2_align
        c.border = header_bdr

    # RKA header (merge baris 3)
    rka_tahun = now.year
    rka_bulan_terakhir_idx = 0
    if periode_list:
        lbl = periode_list[-1]
        parts = lbl.split('-')
        if len(parts) >= 2:
            thn_str = parts[-1]
            bln_str = parts[0].lower()
            try:
                rka_tahun = int("20" + thn_str) if len(thn_str) == 2 else int(thn_str)
            except: pass
            
            from ui.input_rka_widget import MONTHS_MAP
            for i, (k, v) in enumerate(MONTHS_MAP.items()):
                if k in bln_str:
                    rka_bulan_terakhir_idx = i
                    break

    # RKA Header Row 3
    c = ws.cell(row=HDR_ROW1, column=COL_RKA_START, value="RKA")
    c.fill = header_fill
    c.font = header_font
    c.alignment = header_align
    c.border = header_bdr
    
    for ci in range(COL_RKA_START + 1, COL_RKA_END + 1):
        c2 = ws.cell(row=HDR_ROW1, column=ci, value="")
        c2.border = header_bdr
        c2.fill = header_fill

    # RKA Header Row 4 (12 Bulan)
    bulan_singkat_rka = ["Januari", "Februari", "Maret", "Apr", "Mei", "Juni", "Juli", "Ags", "Sep", "Okt", "Nov", "Des"]
    for i in range(12):
        c = ws.cell(row=HDR_ROW2, column=COL_RKA_START + i, value=f"{bulan_singkat_rka[i]}-{str(rka_tahun)[-2:]}")
        c.fill = header2_fill
        c.font = header2_font
        c.alignment = header2_align
        c.border = header_bdr

    # Pencapaian RKA header (merge baris 3-4)
    pencap_label = f"{BULAN_SINGKAT[now.month]}-{str(now.year)[-2:]}"
    for r in [HDR_ROW1, HDR_ROW2]:
        c = ws.cell(row=r, column=COL_PENCAP,
                    value="Pencp RKA %" if r == HDR_ROW1 else pencap_label)
        c.fill = header_fill
        c.font = header_font
        c.alignment = header_align
        c.border = header_bdr

    # Growth header grup (baris 3)
    for ci in range(COL_MTD, COL_YTD + 1):
        c = ws.cell(row=HDR_ROW1, column=ci,
                    value="Growth" if ci == COL_MTD else "")
        c.fill = header_fill
        c.font = header_font
        c.alignment = header_align
        c.border = header_bdr

    # Growth detail (baris 4)
    for col, lbl in [(COL_MTD, mtd_label), (COL_DTD, dtd_label),
                     (COL_YOY, yoy_label), (COL_YTD, ytd_label)]:
        c = ws.cell(row=HDR_ROW2, column=col, value=lbl)
        c.fill = header2_fill
        c.font = header2_font
        c.alignment = header2_align
        c.border = header_bdr

    # ── MERGE CELLS ──────────────────────────────────────────────
    ws.merge_cells(start_row=HDR_ROW1, start_column=COL_MATA,
                   end_row=HDR_ROW2, end_column=COL_MATA)

    if n_periode > 1:
        ws.merge_cells(start_row=HDR_ROW1, start_column=COL_POS_START,
                       end_row=HDR_ROW1, end_column=COL_POS_END)

    ws.merge_cells(start_row=HDR_ROW1, start_column=COL_RKA_START,
                   end_row=HDR_ROW1, end_column=COL_RKA_END)

    ws.merge_cells(start_row=HDR_ROW1, start_column=COL_PENCAP,
                   end_row=HDR_ROW2, end_column=COL_PENCAP)

    if COL_YTD > COL_MTD:
        ws.merge_cells(start_row=HDR_ROW1, start_column=COL_MTD,
                       end_row=HDR_ROW1, end_column=COL_YTD)

    ws.row_dimensions[HDR_ROW1].height = 28
    ws.row_dimensions[HDR_ROW2].height = 40

    # ══════════════════════════════════════════════════════════════
    # DATA ROWS (mulai baris 5)
    # ══════════════════════════════════════════════════════════════
    current_row = 5
    data_row_counter = 0  # untuk alternating colors

    # === FETCH RKA DATA ===
    from core.db_manager import get_connection
    conn = get_connection()
    conn.row_factory = __import__('sqlite3').Row
    cur = conn.cursor()
    
    rka_records_by_month = {}
    
    if kc_name == "Total AH Gunsar":
        cols = ["dpk_tabungan", "dpk_giro", "dpk_deposito", "dpk_casa", "korp_giro", "korp_deposito",
                "pinj_mikro", "pinj_small", "pinj_konsumer",
                "sml_mikro", "sml_small", "sml_konsumer", 
                "npl_mikro", "npl_small", "npl_konsumer",
                "rec_mikro", "rec_small", "rec_konsumer"]
        query = f"SELECT bulan, {', '.join(['SUM('+c+') as '+c for c in cols])} FROM rka_data_v2 WHERE tahun = ? GROUP BY bulan"
        cur.execute(query, (rka_tahun,))
        for row in cur.fetchall():
            if row[cols[0]] is not None:
                rka_records_by_month[row["bulan"]] = dict(row)
    else:
        kc_db_name = f"KC Jakarta {kc_name}"
        cur.execute("SELECT * FROM rka_data_v2 WHERE kc = ? AND tahun = ?", (kc_db_name, rka_tahun))
        for row in cur.fetchall():
            rka_records_by_month[row["bulan"]] = dict(row)
            
    conn.close()

    # Mapping logic
    def get_rka_vals(section, label):
        vals = [""] * 12
        if not rka_records_by_month: return vals
        
        mapping = {
            ("Dana Pihak Ketiga", "Dana Pihak Ketiga"): ["dpk_tabungan", "dpk_giro", "dpk_deposito"],
            ("Dana Pihak Ketiga", "Tabungan"): ["dpk_tabungan"],
            ("Dana Pihak Ketiga", "Giro"): ["dpk_giro"],
            ("Dana Pihak Ketiga", "Deposito"): ["dpk_deposito"],
            ("Dana Pihak Ketiga", "CASA"): ["dpk_tabungan", "dpk_giro"],
            
            ("DPK Korporasi", "DPK Korporasi"): ["korp_giro", "korp_deposito"],
            ("DPK Korporasi", "Giro"): ["korp_giro"],
            ("DPK Korporasi", "Deposito"): ["korp_deposito"],
            
            ("Pinjaman", "Pinjaman"): ["pinj_mikro", "pinj_small", "pinj_konsumer"],
            ("Pinjaman", "Mikro"): ["pinj_mikro"],
            ("Pinjaman", "Small"): ["pinj_small"],
            ("Pinjaman", "Konsumer"): ["pinj_konsumer"],
            ("Pinjaman", "Konsumer - KPR"): ["pinj_kons_kpr"],
            ("Pinjaman", "Konsumer - Briguna Ritel"): ["pinj_kons_briguna"],
            
            ("SML", "SML"): ["sml_mikro", "sml_small", "sml_konsumer"],
            ("SML", "SML %"): ["sml_pct"],
            ("SML", "Mikro"): ["sml_mikro"],
            ("SML", "Mikro %"): ["sml_mikro_pct"],
            ("SML", "Small"): ["sml_small"],
            ("SML", "Small %"): ["sml_small_pct"],
            ("SML", "Konsumer"): ["sml_konsumer"],
            ("SML", "Konsumer %"): ["sml_konsumer_pct"],
            ("SML", "Konsumer - KPR"): ["sml_kons_kpr"],
            ("SML", "Konsumer - KPR %"): ["sml_kons_kpr_pct"],
            ("SML", "Konsumer - Briguna Ritel"): ["sml_kons_briguna"],
            ("SML", "Konsumer - Briguna Ritel %"): ["sml_kons_briguna_pct"],
            
            ("NPL", "NPL"): ["npl_mikro", "npl_small", "npl_konsumer"],
            ("NPL", "NPL %"): ["npl_pct"],
            ("NPL", "Mikro"): ["npl_mikro"],
            ("NPL", "Mikro %"): ["npl_mikro_pct"],
            ("NPL", "Small"): ["npl_small"],
            ("NPL", "Small %"): ["npl_small_pct"],
            ("NPL", "Konsumer"): ["npl_konsumer"],
            ("NPL", "Konsumer %"): ["npl_konsumer_pct"],
            ("NPL", "Konsumer - KPR"): ["npl_kons_kpr"],
            ("NPL", "Konsumer - KPR %"): ["npl_kons_kpr_pct"],
            ("NPL", "Konsumer - Briguna Ritel"): ["npl_kons_briguna"],
            ("NPL", "Konsumer - Briguna Ritel %"): ["npl_kons_briguna_pct"],
            
            ("Recovery. EC", "Recovery. EC"): ["rec_mikro", "rec_small", "rec_konsumer"],
            ("Recovery. EC", "Mikro"): ["rec_mikro"],
            ("Recovery. EC", "Small"): ["rec_small"],
            ("Recovery. EC", "Konsumer"): ["rec_konsumer"],
        }
        
        cols = mapping.get((section, label), [])
        if not cols: return vals
        if 'pct' in cols[0] and kc_name == "Total AH Gunsar": return vals
            
        from ui.input_rka_widget import MONTHS_MAP
        for i, (k, v) in enumerate(MONTHS_MAP.items()):
            rec = rka_records_by_month.get(v, {})
            if not rec:
                continue
            if 'pct' in cols[0]:
                if cols[0] == 'sml_pct':
                    num = sum(rec.get(c, 0) or 0 for c in ["sml_mikro", "sml_small", "sml_konsumer"])
                    den = sum(rec.get(c, 0) or 0 for c in ["pinj_mikro", "pinj_small", "pinj_konsumer"])
                elif cols[0] == 'sml_mikro_pct':
                    num = rec.get("sml_mikro", 0) or 0
                    den = rec.get("pinj_mikro", 0) or 0
                elif cols[0] == 'sml_small_pct':
                    num = rec.get("sml_small", 0) or 0
                    den = rec.get("pinj_small", 0) or 0
                elif cols[0] == 'sml_konsumer_pct':
                    num = rec.get("sml_konsumer", 0) or 0
                    den = rec.get("pinj_konsumer", 0) or 0
                elif cols[0] == 'sml_konsumer_kpr_pct':
                    num = rec.get("sml_konsumer_kpr", 0) or 0
                    den = rec.get("pinj_konsumer", 0) or 0
                elif cols[0] == 'sml_konsumer_briguna_pct':
                    num = rec.get("sml_konsumer_briguna", 0) or 0
                    den = rec.get("pinj_konsumer", 0) or 0
                elif cols[0] == 'npl_pct':
                    num = sum(rec.get(c, 0) or 0 for c in ["npl_mikro", "npl_small", "npl_konsumer"])
                    den = sum(rec.get(c, 0) or 0 for c in ["pinj_mikro", "pinj_small", "pinj_konsumer"])
                elif cols[0] == 'npl_mikro_pct':
                    num = rec.get("npl_mikro", 0) or 0
                    den = sum(rec.get(c, 0) or 0 for c in ["pinj_mikro", "pinj_small", "pinj_konsumer"])
                elif cols[0] == 'npl_small_pct':
                    num = rec.get("npl_small", 0) or 0
                    den = sum(rec.get(c, 0) or 0 for c in ["pinj_mikro", "pinj_small", "pinj_konsumer"])
                elif cols[0] == 'npl_konsumer_pct':
                    num = rec.get("npl_konsumer", 0) or 0
                    den = sum(rec.get(c, 0) or 0 for c in ["pinj_mikro", "pinj_small", "pinj_konsumer"])
                elif cols[0] == 'npl_konsumer_kpr_pct':
                    num = rec.get("npl_konsumer_kpr", 0) or 0
                    den = sum(rec.get(c, 0) or 0 for c in ["pinj_mikro", "pinj_small", "pinj_konsumer"])
                elif cols[0] == 'npl_konsumer_briguna_pct':
                    num = rec.get("npl_konsumer_briguna", 0) or 0
                    den = sum(rec.get(c, 0) or 0 for c in ["pinj_mikro", "pinj_small", "pinj_konsumer"])
                else:
                    num, den = 0, 1
                    
                vals[i] = num / den if den != 0 else 0.0
            else:
                vals[i] = sum(rec.get(c, 0) or 0 for c in cols)
        return vals

    current_section = ""

    for row_data in rows_data:
        row_type = row_data.get("row_type", "data")
        label = row_data.get("label", "")
        
        if row_type in ("header", "header_value") or (row_type == "bold" and label in ("SML", "NPL")):
            current_section = label
            
        rka_vals = get_rka_vals(current_section, label)

        if row_type == "separator":
            current_row = _write_separator(ws, current_row, LAST_COL)
        elif row_type == "header":
            current_row = _write_header_seksi(
                ws, current_row, row_data, periode_list,
                COL_MATA, COL_POS_START, COL_RKA_START, COL_RKA_END, COL_PENCAP,
                COL_MTD, COL_DTD, COL_YOY, COL_YTD, LAST_COL)
        elif row_type == "header_value":
            current_row = _write_header_value(
                ws, current_row, row_data, rka_vals, periode_list,
                COL_MATA, COL_POS_START, COL_POS_END, COL_RKA_START, COL_RKA_END, COL_PENCAP,
                COL_MTD, COL_DTD, COL_YOY, COL_YTD, LAST_COL, rka_bulan_terakhir_idx)
        elif row_type == "bold":
            current_row = _write_bold_label(
                ws, current_row, row_data, rka_vals, periode_list,
                COL_MATA, COL_POS_START, COL_POS_END, COL_RKA_START, COL_RKA_END, COL_PENCAP,
                COL_MTD, COL_DTD, COL_YOY, COL_YTD, rka_bulan_terakhir_idx)
        else:
            current_row = _write_data_row(
                ws, current_row, row_data, rka_vals, periode_list,
                COL_MATA, COL_POS_START, COL_POS_END, COL_RKA_START, COL_RKA_END, COL_PENCAP,
                COL_MTD, COL_DTD, COL_YOY, COL_YTD,
                data_row_counter, rka_bulan_terakhir_idx)
            data_row_counter += 1

    # ── CONDITIONAL FORMATTING GROWTH ────────────────────────────
    if current_row > 6:
        for col in [COL_MTD, COL_DTD, COL_YOY, COL_YTD]:
            col_ltr = get_column_letter(col)
            cell_range = f"{col_ltr}5:{col_ltr}{current_row - 1}"
            ws.conditional_formatting.add(
                cell_range,
                CellIsRule(operator="lessThan", formula=["0"],
                           font=Font(color=CLR_NEG_FONT)))
            ws.conditional_formatting.add(
                cell_range,
                CellIsRule(operator="equal", formula=["0"],
                           font=Font(color=CLR_ZERO_FONT)))

    # ── FREEZE PANES ─────────────────────────────────────────────
    ws.freeze_panes = ws.cell(row=5, column=COL_POS_START)

    # ── LEBAR KOLOM ──────────────────────────────────────────────
    ws.column_dimensions["A"].width = 22
    for i in range(n_periode):
        ws.column_dimensions[get_column_letter(COL_POS_START + i)].width = 13
    for i in range(12):
        ws.column_dimensions[get_column_letter(COL_RKA_START + i)].width = 13
    ws.column_dimensions[get_column_letter(COL_PENCAP)].width = 13
    ws.column_dimensions[get_column_letter(COL_MTD)].width = 17
    for col in [COL_DTD, COL_YOY, COL_YTD]:
        ws.column_dimensions[get_column_letter(col)].width = 11


# ────────────────────────────────────────────────────────────────────
# SEPARATOR — baris biru solid, height 6px
# ────────────────────────────────────────────────────────────────────
def _write_separator(ws, row: int, last_col: int) -> int:
    fill_obj = _fill(CLR_SEPARATOR_BG)
    for ci in range(1, last_col + 1):
        c = ws.cell(row=row, column=ci, value="")
        c.fill = fill_obj
    ws.row_dimensions[row].height = 6
    return row + 1


# ────────────────────────────────────────────────────────────────────
# HEADER SEKSI — "Dana Pihak Ketiga", "Pinjaman", "Recovery. EC"
# ────────────────────────────────────────────────────────────────────
def _write_header_seksi(ws, row: int, row_data: dict,
                        periode_list, COL_MATA, COL_POS_START,
                        COL_RKA_START, COL_RKA_END, COL_PENCAP,
                        COL_MTD, COL_DTD, COL_YOY, COL_YTD,
                        LAST_COL) -> int:
    label = row_data.get("label", "")
    fill_obj = _fill(CLR_SECTION_BG)
    font_obj = _font(bold=True, size=11, color="1E293B")
    bdr = _border_bottom_medium()
    alg = _align(h="left", v="center")

    for ci in range(1, LAST_COL + 1):
        c = ws.cell(row=row, column=ci,
                    value=label if ci == COL_MATA else "")
        c.fill = fill_obj
        c.font = font_obj
        c.alignment = alg
        c.border = bdr

    ws.row_dimensions[row].height = 18
    return row + 1


# ────────────────────────────────────────────────────────────────────
# HEADER VALUE — Header seksi yang punya nilai (misal: "Pinjaman")
# ────────────────────────────────────────────────────────────────────
def _write_header_value(ws, row: int, row_data: dict, rka_vals: list,
                        periode_list, COL_MATA, COL_POS_START,
                        COL_POS_END, COL_RKA_START, COL_RKA_END, COL_PENCAP,
                        COL_MTD, COL_DTD, COL_YOY, COL_YTD,
                        LAST_COL, rka_bulan_idx: int) -> int:
    label = row_data.get("label", "")
    values = row_data.get("values", {})
    fill_obj = _fill(CLR_SECTION_BG)
    font_obj = _font(bold=True, size=11, color="1E293B")
    bdr = _border_bottom_medium()
    
    # Mata Anggaran
    c = ws.cell(row=row, column=COL_MATA, value=label)
    c.fill = fill_obj
    c.font = font_obj
    c.alignment = _align(h="left", v="center")
    c.border = bdr

    # Posisi per periode
    for i, lbl in enumerate(periode_list):
        val = values.get(lbl, 0)
        c = ws.cell(row=row, column=COL_POS_START + i)
        try:
            c.value = float(val)
        except (ValueError, TypeError):
            c.value = val if val == "-" else ""

        c.number_format = "#,##0;-#,##0;\"-\""
        c.fill = fill_obj
        c.font = font_obj
        c.alignment = _align(h="right", v="center")
        c.border = bdr

    # RKA
    for i in range(12):
        c = ws.cell(row=row, column=COL_RKA_START + i)
        val = rka_vals[i] if rka_vals else ""
        if val != "":
            try:
                c.value = float(val)
            except (ValueError, TypeError):
                c.value = val
        else:
            c.value = ""
        c.number_format = "#,##0;-#,##0;\"-\""
        c.fill = fill_obj
        c.font = font_obj
        c.alignment = _align(h="right", v="center")
        c.border = bdr

    # Pencapaian RKA
    c = ws.cell(row=row, column=COL_PENCAP)
    label = row_data.get("label", "")
    
    from core.processor import hitung_pencapaian_rka
    val_terbaru = None
    if periode_list:
        terbaru_label = periode_list[-1]
        val_terbaru = row_data.get("values", {}).get(terbaru_label)
        
    val_rka = rka_vals[rka_bulan_idx] if rka_vals and rka_bulan_idx < len(rka_vals) else None
    
    try:
        val_terbaru = float(val_terbaru) if val_terbaru is not None else None
    except:
        val_terbaru = None
        
    try:
        val_rka = float(val_rka) if val_rka is not None else None
    except:
        val_rka = None
        
    pencap = hitung_pencapaian_rka(label, val_terbaru, val_rka)
    
    c.fill = fill_obj
    c.alignment = _align(h="right", v="center")
    c.border = bdr
    c.number_format = "0.00%"
    
    if pencap is not None:
        c.value = pencap
        if pencap >= 1.00:
            c.font = _font(color="16A34A") # Hijau
        elif pencap >= 0.80:
            c.font = _font(color="D97706") # Kuning
        else:
            c.font = _font(color="DC2626") # Merah
    else:
        c.value = ""
        c.font = font_obj

    # Growth (kalau ada)
    is_pct = '%' in label
    for col, key in [(COL_MTD, "mtd"), (COL_DTD, "dtd"),
                     (COL_YOY, "yoy"), (COL_YTD, "ytd")]:
        val = row_data.get(key)
        c = ws.cell(row=row, column=col)
        c.fill = fill_obj
        c.alignment = _align(h="right", v="center")
        c.border = bdr
        
        if val is None or val == "":
            c.value = ""
            c.font = font_obj
        else:
            try:
                num_val = float(val)
                c.value = num_val
                if is_pct:
                    c.number_format = "0.00%"
                else:
                    c.number_format = "#,##0;-#,##0;\"-\""
                    
                if num_val < 0:
                    c.font = _font(color="DC2626") # Merah
                else:
                    c.font = _font(color="1E293B") # Hitam
            except (ValueError, TypeError):
                c.value = val if val == "-" else ""
                c.font = font_obj

    ws.row_dimensions[row].height = 18
    return row + 1




# ────────────────────────────────────────────────────────────────────
# BOLD LABEL — "SML", "SML %", "NPL", "NPL %", "CASA"
# ────────────────────────────────────────────────────────────────────
def _write_bold_label(ws, row: int, row_data: dict, rka_vals: list,
                      periode_list, COL_MATA, COL_POS_START,
                      COL_POS_END, COL_RKA_START, COL_RKA_END, COL_PENCAP,
                      COL_MTD, COL_DTD, COL_YOY, COL_YTD, rka_bulan_idx: int) -> int:
    label = row_data.get("label", "")
    values = row_data.get("values", {})
    fill_obj = _fill(CLR_BOLD_BG)
    font_obj = _font(bold=True, size=10, color="1E293B")
    bdr = _border()

    is_pct = '%' in label

    # Mata Anggaran
    c = ws.cell(row=row, column=COL_MATA, value=label)
    c.fill = fill_obj
    c.font = font_obj
    c.alignment = _align(h="left", v="center")
    c.border = bdr

    # Posisi per periode
    for i, lbl in enumerate(periode_list):
        val = values.get(lbl, 0)
        c = ws.cell(row=row, column=COL_POS_START + i)
        try:
            c.value = float(val)
        except (ValueError, TypeError):
            c.value = val if val == "-" else ""

        c.number_format = '0.00%' if is_pct else "#,##0;-#,##0;\"-\""
        c.fill = fill_obj
        c.font = font_obj
        c.alignment = _align(h="right", v="center")
        c.border = bdr

    # RKA
    for i in range(12):
        c = ws.cell(row=row, column=COL_RKA_START + i)
        val = rka_vals[i] if rka_vals else ""
        if val != "":
            try:
                c.value = float(val)
            except (ValueError, TypeError):
                c.value = val
        else:
            c.value = ""
        c.number_format = '0.00%' if is_pct else "#,##0;-#,##0;\"-\""
        c.fill = _fill(CLR_RKA_BG)
        c.font = font_obj
        c.alignment = _align(h="right", v="center")
        c.border = bdr

    # Pencapaian RKA
    c = ws.cell(row=row, column=COL_PENCAP)
    
    from core.processor import hitung_pencapaian_rka
    val_terbaru = None
    if periode_list:
        terbaru_label = periode_list[-1]
        val_terbaru = row_data.get("values", {}).get(terbaru_label)
        
    val_rka = rka_vals[rka_bulan_idx] if rka_vals and rka_bulan_idx < len(rka_vals) else None
    
    try:
        val_terbaru = float(val_terbaru) if val_terbaru is not None else None
    except:
        val_terbaru = None
        
    try:
        val_rka = float(val_rka) if val_rka is not None else None
    except:
        val_rka = None
        
    pencap = hitung_pencapaian_rka(label, val_terbaru, val_rka)
    
    c.fill = _fill(CLR_PENCAP_BG)
    c.alignment = _align(h="right", v="center")
    c.border = bdr
    c.number_format = "0.00%"
    
    if pencap is not None:
        c.value = pencap
        if pencap >= 1.00:
            c.font = _font(bold=True, color="16A34A") # Hijau
        elif pencap >= 0.80:
            c.font = _font(bold=True, color="D97706") # Kuning
        else:
            c.font = _font(bold=True, color="DC2626") # Merah
    else:
        c.value = ""
        c.font = font_obj

    # Growth
    for col, key in [(COL_MTD, "mtd"), (COL_DTD, "dtd"),
                     (COL_YOY, "yoy"), (COL_YTD, "ytd")]:
        val = row_data.get(key)
        c = ws.cell(row=row, column=col)
        c.fill = fill_obj
        c.alignment = _align(h="right", v="center")
        c.border = bdr
        
        if val is None or val == "":
            c.value = ""
            c.font = font_obj
        else:
            try:
                num_val = float(val)
                c.value = num_val
                if is_pct:
                    c.number_format = "0.00%"
                else:
                    c.number_format = "#,##0;-#,##0;\"-\""
                    
                if num_val < 0:
                    c.font = _font(bold=True, color="DC2626") # Merah
                else:
                    c.font = font_obj # Hitam Bold
            except (ValueError, TypeError):
                c.value = val if val == "-" else ""
                c.font = font_obj

    ws.row_dimensions[row].height = 18
    return row + 1


# ────────────────────────────────────────────────────────────────────
# DATA ROW — baris data biasa
# ────────────────────────────────────────────────────────────────────
def _write_data_row(ws, row: int, row_data: dict, rka_vals: list,
                    periode_list, COL_MATA, COL_POS_START,
                    COL_POS_END, COL_RKA_START, COL_RKA_END, COL_PENCAP,
                    COL_MTD, COL_DTD, COL_YOY, COL_YTD,
                    counter: int, rka_bulan_idx: int) -> int:
    label = row_data.get("label", "")
    values = row_data.get("values", {})
    fill_obj = _fill(CLR_ROW_A if counter % 2 == 0 else CLR_ROW_B)
    font_obj = _font(size=10, color=CLR_POS_FONT)
    bdr = _border()

    is_pct = '%' in label

    # Mata Anggaran
    c = ws.cell(row=row, column=COL_MATA, value=label)
    c.fill = fill_obj
    c.font = font_obj
    c.alignment = _align(h="left", v="center", indent=1)
    c.border = bdr

    # Posisi per periode
    for i, lbl in enumerate(periode_list):
        val = values.get(lbl, 0)
        c = ws.cell(row=row, column=COL_POS_START + i)
        try:
            c.value = float(val)
        except (ValueError, TypeError):
            c.value = val if val == "-" else ""

        c.number_format = '0.00%' if is_pct else "#,##0;-#,##0;\"-\""
        c.fill = fill_obj
        c.font = font_obj
        c.alignment = _align(h="right", v="center")
        c.border = bdr

    # RKA
    for i in range(12):
        c = ws.cell(row=row, column=COL_RKA_START + i)
        val = rka_vals[i] if rka_vals else ""
        if val != "":
            try:
                c.value = float(val)
            except (ValueError, TypeError):
                c.value = val
        else:
            c.value = ""
        c.number_format = '0.00%' if is_pct else "#,##0;-#,##0;\"-\""
        c.fill = _fill(CLR_RKA_BG)
        c.font = font_obj
        c.alignment = _align(h="right", v="center")
        c.border = bdr

    # Pencapaian RKA
    c = ws.cell(row=row, column=COL_PENCAP)
    
    from core.processor import hitung_pencapaian_rka
    val_terbaru = None
    if periode_list:
        terbaru_label = periode_list[-1]
        val_terbaru = row_data.get("values", {}).get(terbaru_label)
        
    val_rka = rka_vals[rka_bulan_idx] if rka_vals and rka_bulan_idx < len(rka_vals) else None
    
    try:
        val_terbaru = float(val_terbaru) if val_terbaru is not None else None
    except:
        val_terbaru = None
        
    try:
        val_rka = float(val_rka) if val_rka is not None else None
    except:
        val_rka = None
        
    pencap = hitung_pencapaian_rka(label, val_terbaru, val_rka)
    
    c.fill = _fill(CLR_PENCAP_BG)
    c.alignment = _align(h="right", v="center")
    c.border = bdr
    c.number_format = "0.00%"
    
    if pencap is not None:
        c.value = pencap
        if pencap >= 1.00:
            c.font = _font(color="16A34A") # Hijau
        elif pencap >= 0.80:
            c.font = _font(color="D97706") # Kuning
        else:
            c.font = _font(color="DC2626") # Merah
    else:
        c.value = ""
        c.font = font_obj

    # Growth
    for col, key in [(COL_MTD, "mtd"), (COL_DTD, "dtd"),
                     (COL_YOY, "yoy"), (COL_YTD, "ytd")]:
        val = row_data.get(key)
        c = ws.cell(row=row, column=col)
        c.fill = fill_obj
        c.alignment = _align(h="right", v="center")
        c.border = bdr
        
        if val is None or val == "":
            c.value = ""
            c.font = font_obj
        else:
            try:
                num_val = float(val)
                c.value = num_val
                if is_pct:
                    c.number_format = "0.00%"
                else:
                    c.number_format = "#,##0;-#,##0;\"-\""
                    
                if num_val < 0:
                    c.font = _font(color="DC2626") # Merah
                else:
                    c.font = font_obj # Hitam (or default)
            except (ValueError, TypeError):
                c.value = val if val == "-" else ""
                c.font = font_obj

    ws.row_dimensions[row].height = 18
    return row + 1


# ────────────────────────────────────────────────────────────────────
# UTILITY
# ────────────────────────────────────────────────────────────────────
def get_file_size_str(path: str) -> str:
    p = Path(path)
    if not p.exists():
        return "0 KB"
    sz = p.stat().st_size
    if sz < 1_048_576:
        return f"{sz / 1024:.0f} KB"
    return f"{sz / 1_048_576:.1f} MB"

